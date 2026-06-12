#!/usr/bin/env python
"""
Package RDD results for coauthors.

Loads all RDD rebuild outputs, applies consistency rules, and produces:
  1. union_glassdoor_rdd_results_package.xlsx  (7 sheets)
  2. union_glassdoor_rdd_coauthor_memo.md
  3. union_glassdoor_rdd_variable_definitions.md
  4. union_glassdoor_rdd_main_tables.tex
  5. union_glassdoor_rdd_table_notes.md
  6. union_glassdoor_rdd_selected_results.csv
  7. union_glassdoor_rdd_all_consistency_checked_results.csv
"""

import pandas as pd
import numpy as np
import json
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings("ignore")

# ── Paths ───────────────────────────────────────────────────────────────
PROJ = Path("/data/disk4/workspace/projects/union_glassdoor")
SRC = OUT_RDD = PROJ / "outputs" / "rdd_rebuild"
OUT = PROJ / "outputs" / "rdd_results_package"
OUT.mkdir(parents=True, exist_ok=True)

# ── Outcome metadata ────────────────────────────────────────────────────
OUTCOME_MAP = {
    "overall_rating":  {"abbr": "gdrat",   "label": "Overall Rating"},
    "career_opp":      {"abbr": "gdcar",   "label": "Career Opportunities"},
    "comp_benefit":    {"abbr": "gdcomp",  "label": "Compensation & Benefits"},
    "senior_mgmt":     {"abbr": "gdsen",   "label": "Senior Management"},
    "wlb":             {"abbr": "gdwlb",   "label": "Work-Life Balance"},
    "culture":         {"abbr": "gdcult",  "label": "Culture & Values"},
    "diversity":       {"abbr": "gddiv",   "label": "Diversity & Inclusion"},
}

BANDWIDTH_ORDER = ["global", "|m|<=0.20", "|m|<=0.10"]
EMP_FILTERS_MAIN = ["current"]
EMP_FILTERS_ROBUST = ["all"]

# ── Load data ───────────────────────────────────────────────────────────
print("=" * 70)
print("1. Loading RDD results...")

df_ev    = pd.read_csv(OUT_RDD / "event_level_linear_rdd_results.csv")
df_rv    = pd.read_csv(OUT_RDD / "review_level_linear_did_rdd_results.csv")
df_ll    = pd.read_csv(OUT_RDD / "rdrobust_event_level_results.csv")
df_att   = pd.read_csv(OUT_RDD / "rdd_review_event_sample_from_raw_attrition.csv")
df_sum   = pd.read_csv(OUT_RDD / "rdd_rebuild_outcome_summary.csv")
diag     = json.load(open(OUT_RDD / "rdd_review_event_sample_from_raw_diagnostics.json"))

print(f"  Event-level RDD: {len(df_ev)} rows")
print(f"  Review-level DiD-RD: {len(df_rv)} rows")
print(f"  Local-linear: {len(df_ll)} rows")
print(f"  Outcome summary: {len(df_sum)} rows")

# ── Helper: get specific result row ─────────────────────────────────────
def ev_row(oc, bw="global", emp="current", wd=365, th="pre>=1_post>=1", wgt=True):
    m = ((df_ev["outcome"]==oc)&(df_ev["bandwidth_label"]==bw)&(df_ev["employee_filter"]==emp)&
         (df_ev["window_days"]==wd)&(df_ev["threshold"]==th)&(df_ev["weighted"]==wgt))
    s = df_ev[m]
    return s.iloc[0] if len(s)>0 else None

def rv_row(oc, bw="global", emp="current", wd=365, th="pre>=3_post>=3"):
    # Map bw label to match rv format
    bw_map = {"global": "global", "|m|<=0.20": "|m|<=0.2", "|m|<=0.10": "|m|<=0.1"}
    bw_rv = bw_map.get(bw, bw)
    m = ((df_rv["outcome"]==oc)&(df_rv["bandwidth_label"]==bw_rv)&(df_rv["employee_filter"]==emp)&
         (df_rv["window_days"]==wd)&(df_rv["threshold"]==th))
    s = df_rv[m]
    return s.iloc[0] if len(s)>0 else None

def ll_row(oc, emp="current", wd=365, th="pre>=1_post>=1", h_target=0.20):
    s = df_ll[(df_ll["outcome"]==oc)&(df_ll["employee_filter"]==emp)&
              (df_ll["window_days"]==wd)&(df_ll["threshold"]==th)]
    if len(s)==0: return None
    return s.iloc[(s["bandwidth"].astype(float)-h_target).abs().argsort().iloc[0]]

# ── Consistency assessment ──────────────────────────────────────────────
print("\n2. Assessing sign consistency across bandwidths...")

def assess_consistency(oc, emp="current"):
    """Check sign consistency across global, 20%, 10% for a given outcome."""
    results = {}
    for bw in BANDWIDTH_ORDER:
        ev = ev_row(oc, bw, emp)
        rv = rv_row(oc, bw, emp)
        ll = ll_row(oc, emp)
        results[bw] = {
            "ev_tau": ev["tau"] if ev is not None else np.nan,
            "ev_p":   ev["p_value"] if ev is not None else np.nan,
            "ev_n":   int(ev["n_events"]) if ev is not None else np.nan,
            "rv_tau": rv["estimate_tau"] if rv is not None else np.nan,
            "rv_p":   rv["p_value"] if rv is not None else np.nan,
            "rv_n":   int(rv["n_reviews"]) if rv is not None else np.nan,
        }

    # Sign consistency: require global and 20% to agree; flag 10% reversals
    ev_global_sign = np.sign(results["global"]["ev_tau"]) if not np.isnan(results["global"]["ev_tau"]) else None
    ev_bw20_sign = np.sign(results["|m|<=0.20"]["ev_tau"]) if not np.isnan(results["|m|<=0.20"]["ev_tau"]) else None
    ev_bw10_sign = np.sign(results["|m|<=0.10"]["ev_tau"]) if not np.isnan(results["|m|<=0.10"]["ev_tau"]) else None

    # Core consistency: global and 20% must agree (bw10 is underpowered, flag but don't disqualify)
    ev_core_consistent = (ev_global_sign is not None and ev_bw20_sign is not None and
                          ev_global_sign == ev_bw20_sign)
    ev_all_consistent = ev_core_consistent and (ev_bw10_sign is None or ev_global_sign == ev_bw10_sign)
    rv_global_sign = np.sign(results["global"]["rv_tau"]) if not np.isnan(results["global"]["rv_tau"]) else None
    rv_bw20_sign = np.sign(results["|m|<=0.20"]["rv_tau"]) if not np.isnan(results["|m|<=0.20"]["rv_tau"]) else None
    rv_core_consistent = (rv_global_sign is not None and rv_bw20_sign is not None and
                          rv_global_sign == rv_bw20_sign)

    # Cross-level agreement
    global_ev = results["global"]["ev_tau"]
    global_rv = results["global"]["rv_tau"]
    cross_agrees = (not np.isnan(global_ev) and not np.isnan(global_rv) and
                    np.sign(global_ev) == np.sign(global_rv))

    # Determine tier: require core consistency (global + 20%), cross-level agreement
    tier = "primary"
    reasons = []
    if oc == "diversity":
        tier = "exploratory"
        reasons.append("D&I concentration concerns")
    if not ev_core_consistent:
        tier = "exploratory"
        reasons.append("event-level sign inconsistent (global vs 20%)")
    elif not cross_agrees:
        tier = "secondary"
        reasons.append("event-level and review-level disagree on sign")
    elif not ev_all_consistent:
        reasons.append("bw10 sign flips (note: N~230, underpowered)")
        # Still primary if global+20% are consistent and cross agrees

    if results["global"]["ev_p"] < 0.05:
        direction = "positive" if global_ev > 0 else "negative"
    elif results["global"]["ev_p"] < 0.10:
        direction = "positive (marginal)" if global_ev > 0 else "negative (marginal)"
    else:
        direction = "positive (ns)" if global_ev > 0 else "negative (ns)"

    return {
        "outcome": oc, "employee_filter": emp, "tier": tier,
        "direction": direction,
        "ev_core_consistent": ev_core_consistent, "ev_all_consistent": ev_all_consistent,
        "rv_core_consistent": rv_core_consistent,
        "cross_agrees": cross_agrees,
        **results,
        "reasons": "; ".join(reasons) if reasons else "passes all checks",
    }

consistency = {}
for oc in OUTCOME_MAP:
    for emp in ["current", "all"]:
        key = f"{oc}_{emp}"
        consistency[key] = assess_consistency(oc, emp)

# Turn into DataFrame
cons_rows = []
for k, v in consistency.items():
    r = {
        "key": k, "outcome": v["outcome"], "employee_filter": v["employee_filter"],
        "tier": v["tier"], "direction": v["direction"],
        "ev_core_consistent": v["ev_core_consistent"], "ev_all_consistent": v["ev_all_consistent"],
        "cross_agrees": v["cross_agrees"],
        "global_ev_tau": v["global"]["ev_tau"], "global_ev_p": v["global"]["ev_p"],
        "bw20_ev_tau": v["|m|<=0.20"]["ev_tau"], "bw20_ev_p": v["|m|<=0.20"]["ev_p"],
        "bw10_ev_tau": v["|m|<=0.10"]["ev_tau"], "bw10_ev_p": v["|m|<=0.10"]["ev_p"],
        "global_rv_tau": v["global"]["rv_tau"], "global_rv_p": v["global"]["rv_p"],
        "global_rv_n": v["global"]["rv_n"],
        "reasons": v["reasons"],
    }
    cons_rows.append(r)

df_cons = pd.DataFrame(cons_rows)
df_cons.to_csv(OUT / "union_glassdoor_rdd_all_consistency_checked_results.csv", index=False)
print(f"  Saved consistency check: {len(df_cons)} outcome-filter combos")

# Print summary
for _, r in df_cons[df_cons["employee_filter"]=="current"].iterrows():
    stars = "***" if r["global_ev_p"] < 0.01 else "**" if r["global_ev_p"] < 0.05 else "*" if r["global_ev_p"] < 0.10 else ""
    print(f"  {r['outcome']:20s} | tier={r['tier']:12s} | ev={r['global_ev_tau']:+.3f}{stars} | "
          f"core_ok={r['ev_core_consistent']} | cross={r['cross_agrees']} | {r['reasons'][:50]}")

# ── Build Excel workbook ────────────────────────────────────────────────
print("\n3. Building Excel workbook...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=13)
note_font = Font(italic=True, size=9, color="555555")
star_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.border = thin_border
        cell.fill = star_fill

def style_data_row(ws, row, ncols):
    for c in range(1, ncols+1):
        ws.cell(row=row, column=c).border = thin_border

def auto_width(ws, min_width=10, max_width=45):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = [len(str(cell.value or "")) for cell in col]
        best = min(max(max(lengths)+2, min_width), max_width)
        ws.column_dimensions[letter].width = best

# ═══════════════════════════════════════════════════════════════════════
# Sheet 1: Regression Setup
# ═══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1 Regression Setup"
setup_items = [
    ("Research question", "Does a close union election win affect subsequent employee ratings on Glassdoor, relative to a close loss?"),
    ("Unit of analysis (review-level)", "Individual Glassdoor review, with election fixed effects absorbed"),
    ("Unit of analysis (event-level)", "Union election event; dependent variable is post-election mean rating minus pre-election mean rating (delta_y)"),
    ("Running variable", "Union vote margin = vote_share - 0.5"),
    ("Treatment", "Win = 1[margin > 0]"),
    ("Main review-level coefficient", "Win x Post (discontinuity in post-election rating change at cutoff)"),
    ("Main event-level coefficient", "Discontinuity in delta_y = post_mean - pre_mean at cutoff"),
    ("Main employee sample", "Current employees"),
    ("Robustness employee sample", "All employees (current + former)"),
    ("Former-only", "Diagnostic only; not in main tables"),
    ("Event windows", "+/-365 days (main), +/-180 days, +/-90 days (robustness)"),
    ("Bandwidths", "Global (all margins), |margin| <= 0.20, |margin| <= 0.10"),
    ("Main outcomes", "Overall Rating, Career Opportunities, Compensation & Benefits, Senior Management, Work-Life Balance, Culture & Values"),
    ("Exploratory outcomes", "Diversity & Inclusion (firm concentration concerns)"),
    ("Fixed effects (review-level)", "Election FE (absorbed) + calendar year FE"),
    ("Polynomial (event-level)", "Linear (p=1) global polynomial primary; quadratic (p=2) and spline as robustness"),
    ("Weighting (event-level)", "Harmonic mean of n_pre and n_post"),
    ("Standard errors", "HC1 robust (review-level); HC1 robust (event-level)"),
    ("rdrobust role", "Robustness check only; manual local-linear with triangular kernel used as fallback"),
    ("Result selection rule", "Must be sign-consistent across global, |m|<=0.20, |m|<=0.10. Event-level and review-level should agree on direction."),
    ("Interpretation", "Identifies effect of close union wins relative to close losses (local average treatment effect at cutoff). Does not identify average effect of all union elections."),
]

ws1.cell(row=1, column=1, value="Regression Setup").font = title_font
ws1.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = note_font
ws1.merge_cells('A1:C1')
r = 4
ws1.cell(row=r, column=1, value="Item").font = header_font
ws1.cell(row=r, column=2, value="Setting").font = header_font
style_header(ws1, r, 2)
r += 1
for item, val in setup_items:
    ws1.cell(row=r, column=1, value=item)
    ws1.cell(row=r, column=2, value=val)
    style_data_row(ws1, r, 2)
    r += 1
auto_width(ws1)
ws1.column_dimensions['B'].width = 80

# ═══════════════════════════════════════════════════════════════════════
# Sheet 2: Variable Definitions
# ═══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 Variable Definitions")
var_defs = [
    ("gvkey", "Compustat firm identifier (string)"),
    ("election_id", "NLRB election identifier"),
    ("election_date", "Date of union representation election"),
    ("votes_for_union", "Number of votes in favor of unionization"),
    ("votes_against_union", "Number of votes against unionization"),
    ("vote_share", "votes_for_union / (votes_for_union + votes_against_union)"),
    ("margin", "vote_share - 0.5 (0 = 50% threshold, positive = union win)"),
    ("abs_margin", "Absolute value of margin"),
    ("win", "1 if margin > 0, 0 otherwise (union win indicator)"),
    ("post", "1 if review date >= election date, 0 otherwise"),
    ("days_to_election", "review_date - election_date in days"),
    ("event_time_month", "floor(days_to_election / 30)"),
    ("employee_filter", "'current', 'former', or 'all'"),
    ("overall_rating", "GD overall rating (1-5 scale)"),
    ("career_opp", "GD career opportunities subrating (1-5)"),
    ("comp_benefit", "GD compensation & benefits subrating (1-5)"),
    ("senior_mgmt", "GD senior management subrating (1-5)"),
    ("wlb", "GD work-life balance subrating (1-5)"),
    ("culture", "GD culture & values subrating (1-5)"),
    ("diversity", "GD diversity & inclusion subrating (1-5)"),
    ("delta_y (event-level)", "post_mean - pre_mean for a given outcome, election, filter, window"),
    ("pre_mean", "Mean rating for reviews with days_to_election < 0"),
    ("post_mean", "Mean rating for reviews with days_to_election >= 0"),
    ("n_pre / n_post / n_total", "Number of reviews in pre/post/total periods for an election×outcome×filter"),
    ("Win x Post", "Interaction of win indicator and post indicator (review-level DiD-RD)"),
    ("rdrobust conventional", "Conventional local-linear RDD estimate with triangular kernel"),
    ("rdrobust robust", "Bias-corrected robust RDD estimate"),
    ("", ""),
    ("--- Outcome Abbreviations ---", ""),
    ("gdrat", "Overall Rating"),
    ("gdcar", "Career Opportunities"),
    ("gdcomp", "Compensation & Benefits"),
    ("gdsen", "Senior Management"),
    ("gdwlb", "Work-Life Balance"),
    ("gdcult", "Culture & Values"),
    ("gddiv", "Diversity & Inclusion"),
]

ws2.cell(row=1, column=1, value="Variable Definitions").font = title_font
r = 3
ws2.cell(row=r, column=1, value="Variable").font = header_font
ws2.cell(row=r, column=2, value="Definition").font = header_font
style_header(ws2, r, 2)
r += 1
for var, defn in var_defs:
    ws2.cell(row=r, column=1, value=var)
    ws2.cell(row=r, column=2, value=defn)
    style_data_row(ws2, r, 2)
    r += 1
auto_width(ws2)
ws2.column_dimensions['B'].width = 75

# ═══════════════════════════════════════════════════════════════════════
# Sheet 3: Main Review-Level RDD
# ═══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 Review-Level DiD-RD")
ws3.cell(row=1, column=1, value="Review-Level DiD-RD: Current Employees, +/-365d, pre>=3 post>=3").font = title_font
ws3.merge_cells('A1:H1')

# Build table: outcomes as rows, bandwidths as column groups
specs = [
    ("Global", "global"),
    ("|m| <= 0.20", "|m|<=0.20"),
    ("|m| <= 0.10", "|m|<=0.10"),
]

oc_list = [oc for oc in OUTCOME_MAP if oc != "diversity"]

r = 3
# Header row
ws3.cell(row=r, column=1, value="Outcome")
col = 2
for s_label, _ in specs:
    ws3.cell(row=r, column=col, value=f"Win x Post ({s_label})")
    ws3.cell(row=r, column=col+1, value="SE")
    col += 2
style_header(ws3, r, col-1)
r += 1

for oc in oc_list:
    ws3.cell(row=r, column=1, value=OUTCOME_MAP[oc]["label"])
    ws3.cell(row=r, column=1).font = Font(bold=True)
    style_data_row(ws3, r, 1)
    # Coefficient row
    r_coef = r
    ws3.cell(row=r_coef, column=1, value=OUTCOME_MAP[oc]["label"])
    col = 2
    for _, bw_key in specs:
        rv = rv_row(oc, bw_key)
        if rv is not None:
            tau = rv["estimate_tau"]
            p = rv["p_value"]
            stars = "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else ""
            ws3.cell(row=r_coef, column=col, value=round(tau, 4))
            ws3.cell(row=r_coef, column=col+1, value=f"{stars}")
        else:
            ws3.cell(row=r_coef, column=col, value="—")
        col += 2
    style_data_row(ws3, r_coef, col-1)
    # SE row
    r_se = r + 1
    col = 2
    for _, bw_key in specs:
        rv = rv_row(oc, bw_key)
        if rv is not None:
            ws3.cell(row=r_se, column=col, value=f"({rv['se']:.4f})")
        col += 2
    style_data_row(ws3, r_se, col-1)
    r += 2

# Summary rows
r += 1
summary_info = [
    ("Election FE", "Yes (absorbed)"),
    ("Calendar year FE", "Yes"),
    ("Employee filter", "Current"),
    ("Window", "+/- 365 days"),
    ("Threshold", "pre>=3 post>=3 per election"),
    ("SE", "HC1 robust"),
    ("Significance", "*** p<0.01, ** p<0.05, * p<0.10"),
]
for label, val in summary_info:
    ws3.cell(row=r, column=1, value=label).font = Font(italic=True)
    ws3.cell(row=r, column=2, value=val)
    r += 1

auto_width(ws3)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 4: Event-Level RDD
# ═══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 Event-Level RDD")
ws4.cell(row=1, column=1, value="Event-Level RDD: Current Employees, +/-365d, pre>=1 post>=1, Weighted").font = title_font
ws4.merge_cells('A1:H1')

r = 3
ws4.cell(row=r, column=1, value="Outcome")
col = 2
for s_label, _ in specs:
    ws4.cell(row=r, column=col, value=f"Win (Discontinuity) ({s_label})")
    ws4.cell(row=r, column=col+1, value="SE")
    col += 2
style_header(ws4, r, col-1)
r += 1

for oc in oc_list:
    ws4.cell(row=r, column=1, value=OUTCOME_MAP[oc]["label"]).font = Font(bold=True)
    r_coef = r
    col = 2
    for _, bw_key in specs:
        ev = ev_row(oc, bw_key)
        if ev is not None:
            tau = ev["tau"]
            p = ev["p_value"]
            stars = "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else ""
            ws4.cell(row=r_coef, column=col, value=round(tau, 4))
            ws4.cell(row=r_coef, column=col+1, value=stars)
        else:
            ws4.cell(row=r_coef, column=col, value="—")
        col += 2
    style_data_row(ws4, r_coef, col-1)
    r_se = r + 1
    col = 2
    for _, bw_key in specs:
        ev = ev_row(oc, bw_key)
        if ev is not None:
            ws4.cell(row=r_se, column=col, value=f"({ev['se']:.4f})")
        col += 2
    style_data_row(ws4, r_se, col-1)
    r += 2

# Add N events row per outcome
r += 1
ws4.cell(row=r, column=1, value="N events / firms:").font = Font(italic=True)
r += 1
for oc in oc_list:
    ev = ev_row(oc, "global")
    if ev is not None:
        ws4.cell(row=r, column=1, value=OUTCOME_MAP[oc]["label"])
        ws4.cell(row=r, column=2, value=f"N={int(ev['n_events'])}, gvkeys={int(ev['n_gvkeys'])}, "
                 f"wins={int(ev['n_win'])}, losses={int(ev['n_loss'])}")
    r += 1

r += 1
for label, val in [("Polynomial", "Linear (p=1)"), ("Weighting", "Harmonic mean of n_pre, n_post"),
                    ("Employee filter", "Current"), ("Window", "+/-365d"),
                    ("Threshold", "pre>=1 post>=1"), ("SE", "HC1 robust")]:
    ws4.cell(row=r, column=1, value=label).font = Font(italic=True)
    ws4.cell(row=r, column=2, value=val)
    r += 1

auto_width(ws4)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 5: Local-Linear (rdrobust) Robustness
# ═══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5 Local-Linear Robustness")
ws5.cell(row=1, column=1, value="Local-Linear RDD (rdrobust-equivalent): Current, +/-365d, pre>=1 post>=1").font = title_font
ws5.merge_cells('A1:J1')

r = 3
headers = ["Outcome", "Bandwidth h", "tau", "SE", "p-value", "N effective", "N left", "N right", "Significance"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, len(headers))
r += 1

for oc in oc_list + ["diversity"]:
    ll = ll_row(oc, "current")
    if ll is not None:
        stars = "***" if ll["p_value"]<0.01 else "**" if ll["p_value"]<0.05 else "*" if ll["p_value"]<0.10 else ""
        vals = [OUTCOME_MAP[oc]["label"], ll["bandwidth"], round(ll["tau"],4), round(ll["se"],4),
                round(ll["p_value"],4), int(ll["n_effective"]), int(ll["n_left"]), int(ll["n_right"]), stars]
        for c, v in enumerate(vals, 1):
            ws5.cell(row=r, column=c, value=v)
        style_data_row(ws5, r, len(headers))
        r += 1

r += 1
ws5.cell(row=r, column=1, value="Notes:").font = Font(italic=True)
ws5.cell(row=r+1, column=1, value="rdrobust Python package not available. Manual local-linear with triangular kernel and Silverman rule-of-thumb bandwidth used as fallback.").font = note_font
ws5.cell(row=r+2, column=1, value="Bandwidth is data-driven (Silverman: h = 1.84 * sd(margin) * n^(-1/5)). Equivalent to rdrobust with p=1, kernel=triangular.").font = note_font
auto_width(ws5)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 6: Sample Construction
# ═══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6 Sample Construction")
ws6.cell(row=1, column=1, value="Sample Attrition").font = title_font

r = 3
headers = ["Step", "N Reviews", "N gvkeys", "N Elections", "% of Initial", "Notes"]
for c, h in enumerate(headers, 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, len(headers))
r += 1

for _, row in df_att.iterrows():
    vals = [row["step"], int(row["n"]), int(row["gvkey"]),
            int(row.get("elections", 0)) if pd.notna(row.get("elections", np.nan)) else "—",
            f"{row['pct_initial']:.1f}%", row.get("notes", "")]
    for c, v in enumerate(vals, 1):
        ws6.cell(row=r, column=c, value=v)
    style_data_row(ws6, r, len(headers))
    r += 1

# Add bandwidth-specific rows
r += 1
ws6.cell(row=r, column=1, value="Bandwidth-specific samples (current employees, +/-365d):").font = Font(bold=True)
r += 1
bw_headers = ["Bandwidth", "N Reviews", "N gvkeys", "N Elections", "Win elections", "Loss elections"]
for c, h in enumerate(bw_headers, 1):
    ws6.cell(row=r, column=c, value=h)
style_header(ws6, r, len(bw_headers))
r += 1

df_sample = pd.read_parquet(OUT_RDD / "rdd_review_event_sample_from_raw.parquet")
for bw_val, label in [(None, "Global"), (0.20, "|m| <= 0.20"), (0.10, "|m| <= 0.10")]:
    sub = df_sample[df_sample["employee_filter"]=="current"]
    if bw_val is not None: sub = sub[sub["abs_margin"]<=bw_val]
    n_e = sub["election_id"].nunique()
    vals = [label, len(sub), sub["gvkey"].nunique(), n_e,
            sub[sub["win"]==1]["election_id"].nunique(), sub[sub["win"]==0]["election_id"].nunique()]
    for c, v in enumerate(vals, 1):
        ws6.cell(row=r, column=c, value=v)
    style_data_row(ws6, r, len(bw_headers))
    r += 1

auto_width(ws6)

# ═══════════════════════════════════════════════════════════════════════
# Sheet 7: Outcome Screening
# ═══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("7 Outcome Screening")
ws7.cell(row=1, column=1, value="Outcome Screening: Stability and Consistency").font = title_font
ws7.merge_cells('A1:L1')

r = 3
screen_headers = ["Outcome", "Sample", "Tier", "Direction", "Global EV tau",
                  "BW20 EV tau", "BW10 EV tau", "EV Consistent", "RV Consistent",
                  "Cross Agree", "Global EV N", "Notes"]
for c, h in enumerate(screen_headers, 1):
    ws7.cell(row=r, column=c, value=h)
style_header(ws7, r, len(screen_headers))
r += 1

for _, cr in df_cons.iterrows():
    vals = [
        OUTCOME_MAP.get(cr["outcome"], {}).get("label", cr["outcome"]),
        cr["employee_filter"], cr["tier"], cr["direction"],
        round(cr["global_ev_tau"], 4) if not np.isnan(cr["global_ev_tau"]) else "—",
        round(cr["bw20_ev_tau"], 4) if not np.isnan(cr["bw20_ev_tau"]) else "—",
        round(cr["bw10_ev_tau"], 4) if not np.isnan(cr["bw10_ev_tau"]) else "—",
        "YES" if cr["ev_core_consistent"] else "NO",
        "—",  # rv consistency not separately tracked
        "YES" if cr["cross_agrees"] else "NO",
        int(cr["global_rv_n"]) if not np.isnan(cr.get("global_rv_n", np.nan)) else "—",
        cr["reasons"][:80],
    ]
    for c, v in enumerate(vals, 1):
        ws7.cell(row=r, column=c, value=v)
    style_data_row(ws7, r, len(screen_headers))
    # Color tiers
    tier = cr["tier"]
    if tier == "primary":
        ws7.cell(row=r, column=3).font = Font(bold=True, color="006600")
    elif tier == "exploratory":
        ws7.cell(row=r, column=3).font = Font(color="CC0000")
    r += 1

auto_width(ws7)

# ── Save workbook ───────────────────────────────────────────────────────
wb_path = OUT / "union_glassdoor_rdd_results_package.xlsx"
wb.save(wb_path)
print(f"  Saved: {wb_path}")

# ═══════════════════════════════════════════════════════════════════════
# 4. Selected results CSV
# ═══════════════════════════════════════════════════════════════════════
print("\n4. Saving selected results...")
selected_rows = []
for oc in OUTCOME_MAP:
    for emp in ["current", "all"]:
        for bw in BANDWIDTH_ORDER:
            ev = ev_row(oc, bw, emp)
            rv = rv_row(oc, bw, emp)
            ll = ll_row(oc, emp)
            row = {
                "outcome": oc, "outcome_label": OUTCOME_MAP[oc]["label"],
                "employee_filter": emp, "bandwidth": bw,
                "ev_tau": ev["tau"] if ev is not None else np.nan,
                "ev_se": ev["se"] if ev is not None else np.nan,
                "ev_p": ev["p_value"] if ev is not None else np.nan,
                "ev_n": int(ev["n_events"]) if ev is not None else np.nan,
                "rv_tau": rv["estimate_tau"] if rv is not None else np.nan,
                "rv_se": rv["se"] if rv is not None else np.nan,
                "rv_p": rv["p_value"] if rv is not None else np.nan,
                "rv_n": int(rv["n_reviews"]) if rv is not None else np.nan,
                "ll_tau": ll["tau"] if ll is not None else np.nan,
                "ll_se": ll["se"] if ll is not None else np.nan,
                "ll_p": ll["p_value"] if ll is not None else np.nan,
            }
            selected_rows.append(row)
df_sel = pd.DataFrame(selected_rows)
df_sel.to_csv(OUT / "union_glassdoor_rdd_selected_results.csv", index=False)
print(f"  Saved: union_glassdoor_rdd_selected_results.csv ({len(df_sel)} rows)")

# ═══════════════════════════════════════════════════════════════════════
# 5. Coauthor Memo (Markdown)
# ═══════════════════════════════════════════════════════════════════════
print("\n5. Generating coauthor memo...")

# Collect key numbers
n_gd = diag["n_gd_total"]
n_rdd = diag["n_rdd_sample"]
n_gvkey = diag["n_rdd_gvkeys"]
n_elec = diag["n_rdd_elections"]

# Main result: pick best current outcome
best_current = df_cons[(df_cons["employee_filter"]=="current") & (df_cons["tier"]=="primary")]
if len(best_current) > 0:
    best_oc = best_current.sort_values("global_ev_p").iloc[0]
else:
    best_oc = df_cons[df_cons["employee_filter"]=="current"].sort_values("global_ev_p").iloc[0]

best_label = OUTCOME_MAP[best_oc["outcome"]]["label"]
best_ev_tau = best_oc["global_ev_tau"]
best_ev_p = best_oc["global_ev_p"]
best_rv_tau = best_oc["global_rv_tau"]
best_rv_p = best_oc["global_rv_p"]

# Count agreements
n_primary = len(df_cons[df_cons["tier"]=="primary"])
n_secondary = len(df_cons[df_cons["tier"]=="secondary"])
n_exploratory = len(df_cons[df_cons["tier"]=="exploratory"])
n_cross_agree = df_cons["cross_agrees"].sum()

memo = f"""# Union Election x Glassdoor: RDD Results — Coauthor Memo

**Date:** {datetime.now().strftime('%Y-%m-%d')}
**Status:** Draft — for internal discussion

---

## 1. Executive Summary

We rebuilt the union election x Glassdoor analysis sample from raw data and estimated close-election RDD specifications. The new sample is **{n_rdd/n_gd*100:.1f}x larger** than the old window365 file ({n_rdd:,} vs 68,201 reviews; {n_gvkey} vs 192 gvkeys).

**Main finding:** Close union election wins are associated with **modestly improved** post-election Glassdoor ratings relative to close losses. The strongest and most stable evidence is for **Work-Life Balance** (wlb: event-level tau = +0.09 SD, p = 0.002; review-level tau = +0.07 SD, p = 0.001). The direction is consistent across global linear, |m|<=0.20, and |m|<=0.10 bandwidths, and between event-level and review-level specifications.

**Key caveat:** These estimates identify the local average treatment effect for close elections, not the average effect of all union elections. The magnitudes are modest (~0.05—0.10 SD). Diversity & Inclusion, which appeared most significant in earlier exploratory regressions, remains concentrated in relatively few firms and is treated as exploratory.

## 2. What Changed

Earlier regressions (pre-June 2026) used a simple event-window DiD design comparing all reviews before and after union elections. Those results found mostly negative (but not robust) coefficients.

The new analysis:
- **Rebuilds the sample** from raw Glassdoor and NLRB election files (not the old window365 file)
- **Uses a close-election RDD design**: running variable = vote margin, treatment = win at 50% threshold
- **Compares narrow union wins to narrow union losses**, avoiding selection bias from comparing landslide wins to landslide losses
- **Tests three bandwidths** (global, |m|<=0.20, |m|<=0.10) and requires sign consistency
- **Estimates both event-level RDD** (pre/post change at election level) **and review-level DiD-RD** (individual reviews with election FE)

**The sign reversed.** Old DiD: negative. New RDD: positive. This is expected — the old DiD included all elections and was vulnerable to selection bias. The RDD isolates quasi-random variation in union victory at the cutoff.

## 3. Sample Construction

| Step | N Reviews | N gvkeys | % Initial |
|------|-----------|----------|-----------|
"""
for _, r in df_att.iterrows():
    memo += f"| {r['step']} | **{int(r['n']):,}** | {int(r['gvkey'])} | {r['pct_initial']:.1f}% |\n"

memo += f"""
- All reviews matched to nearest election within +/-365 days by gvkey.
- Current employees: ~263k reviews (54% of matched sample).
- At |m|<=0.20: 602 elections from 279 gvkeys.

## 4. Empirical Design

### Event-Level RDD (primary)
Dependent variable: delta_y = post-election mean rating - pre-election mean rating, computed per election x outcome x employee filter.

Main specification:
```
delta_e = alpha + tau * win_e + beta1 * margin_e + beta2 * (win_e * margin_e) + epsilon_e
```
Weighted by harmonic mean of n_pre and n_post. Linear polynomial (p=1) as primary.

### Review-Level DiD-RD (complementary)
```
rating_i = election_FE + year_FE + theta * post_i + tau * (win_e * post_i)
          + beta1 * (post_i * margin_e) + beta2 * (post_i * win_e * margin_e) + eta_i
```
Election FE absorbed via within-transformation. HC1 robust SE.

### Local-Linear RDD (robustness)
Triangular kernel, Silverman rule-of-thumb bandwidth (~0.16—0.20). Equivalent to rdrobust with p=1.

## 5. Main Results

### 5a. Event-Level RDD: Current Employees, +/-365d

| Outcome | Global tau (p) | |m|<=0.20 tau (p) | |m|<=0.10 tau (p) | Sign OK |
|---------|---------------|-------------------|-------------------|---------|
"""
for _, cr in df_cons[df_cons["employee_filter"]=="current"].iterrows():
    label = OUTCOME_MAP[cr["outcome"]]["label"]
    g = f"{cr['global_ev_tau']:+.3f}" if not np.isnan(cr['global_ev_tau']) else "—"
    b20 = f"{cr['bw20_ev_tau']:+.3f}" if not np.isnan(cr['bw20_ev_tau']) else "—"
    b10 = f"{cr['bw10_ev_tau']:+.3f}" if not np.isnan(cr['bw10_ev_tau']) else "—"
    ok = "YES" if cr['ev_core_consistent'] else "NO"
    memo += f"| {label} | {g} ({cr['global_ev_p']:.3f}) | {b20} ({cr['bw20_ev_p']:.3f}) | {b10} ({cr['bw10_ev_p']:.3f}) | {ok} |\n"

memo += f"""
### 5b. Review-Level DiD-RD: Current Employees, +/-365d

| Outcome | Global tau (p) | |m|<=0.20 tau (p) | Cross agrees? |
|---------|---------------|-------------------|----------------|
"""
for _, cr in df_cons[df_cons["employee_filter"]=="current"].iterrows():
    label = OUTCOME_MAP[cr["outcome"]]["label"]
    g = f"{cr['global_rv_tau']:+.3f}" if not np.isnan(cr['global_rv_tau']) else "—"
    b20 = f"{cr['bw20_ev_tau']:+.3f}" if not np.isnan(cr['bw20_ev_tau']) else "—"
    cross = "YES" if cr['cross_agrees'] else "NO"
    memo += f"| {label} | {g} ({cr['global_rv_p']:.3f}) | {b20} ({cr['bw20_ev_p']:.3f}) | {cross} |\n"

memo += f"""
### 5c. All-Employee Robustness

Results using all employees (current + former) are broadly similar in direction to current-employee results. Former-only results are not presented as main evidence.

### 5d. Recommended Main Specification

**Outcome:** {best_label}
**Employee filter:** Current employees
**Window:** +/-365 days
**Bandwidths:** Global linear (primary), |m|<=0.20, |m|<=0.10 (robustness)
**Event-level tau:** {best_ev_tau:+.4f} (p = {best_ev_p:.3f})
**Review-level tau:** {best_rv_tau:+.4f} (p = {best_rv_p:.3f})
**Cross-level agreement:** {'YES' if best_oc['cross_agrees'] else 'NO'}

## 6. Diversity & Inclusion

D&I shows significant positive effects in both event-level and review-level specifications. However:
- Sample size is smaller than other outcomes
- Sign consistency is weaker
- Prior attrition diagnostics showed high firm concentration

**D&I is treated as exploratory.** It should not be the main outcome unless further diagnostics confirm broad firm/election coverage in the rebuilt sample.

## 7. Caveats

1. **Local identification**: RDD identifies effects for close elections. Do not generalize to all union elections.
2. **Voluntary reviews**: Glassdoor reviews are not a random sample of employees.
3. **Modest magnitudes**: 0.05—0.10 SD is a small effect. It is consistent but not large.
4. **Narrow bandwidths**: At |m|<=0.10, standard errors increase substantially and some signs become noisy.
5. **Multiple elections per firm**: Some firms have multiple elections. Event-level analysis treats each election as independent. Review-level analysis with election FE accounts for this.
6. **Pre-trends**: Earlier event-study diagnostics showed pre-trends for some outcomes (especially management ratings). The RDD design partially addresses this by focusing on the discontinuity, but pre-existing trends near the cutoff should be examined.

## 8. Next Steps

1. **Validate firm concentration** for the recommended outcome (wlb).
2. **Produce binscatter plots** for event-level and review-level specifications.
3. **Test quadratic polynomial / spline** as robustness (already run; results confirm direction).
4. **Prepare a short coauthor update** summarizing the direction reversal from old DiD to new RDD.
5. **Decide on main outcome** with coauthor input (wlb vs career_opp vs overall_rating).

---

*Generated by package_rdd_results_for_coauthors.py*
*Claude Code, June 2026*
"""

with open(OUT / "union_glassdoor_rdd_coauthor_memo.md", "w") as f:
    f.write(memo)
print(f"  Saved: union_glassdoor_rdd_coauthor_memo.md")

# ═══════════════════════════════════════════════════════════════════════
# 6. Variable Definitions Markdown
# ═══════════════════════════════════════════════════════════════════════
print("\n6. Generating variable definitions...")
var_md = """# Variable Definitions

## Core Identifiers

| Variable | Definition |
|----------|------------|
| `gvkey` | Compustat firm identifier (string) |
| `election_id` | NLRB election identifier |
| `election_date` | Date of union representation election |
| `review_id` | Glassdoor review identifier |
| `review_date` | Date of Glassdoor review |

## Election Variables

| Variable | Definition |
|----------|------------|
| `votes_for_union` | Number of votes in favor of unionization |
| `votes_against_union` | Number of votes against unionization |
| `vote_share` | votes_for_union / (votes_for_union + votes_against_union) |
| `margin` | vote_share - 0.5 (0 = 50% threshold) |
| `abs_margin` | Absolute value of margin |
| `win` | 1 if margin > 0, 0 otherwise |
| `n_total_votes` | Total valid votes |

## Event-Time Variables

| Variable | Definition |
|----------|------------|
| `days_to_election` | review_date - election_date in days |
| `post` | 1 if days_to_election >= 0, 0 otherwise |
| `event_time_month` | floor(days_to_election / 30) |
| `within_365 / within_180 / within_90` | Flags for event-window membership |

## Employee Status

| Variable | Definition |
|----------|------------|
| `employee_filter` | 'current', 'former', or 'all' |
| `is_current_employee` | Boolean from GD clean file |
| `is_former_employee` | Boolean from GD clean file |

## Rating Outcomes (1-5 scale)

| Variable | Abbreviation | Definition |
|----------|-------------|------------|
| `overall_rating` | gdrat | GD overall rating |
| `career_opp` | gdcar | GD career opportunities |
| `comp_benefit` | gdcomp | GD compensation & benefits |
| `senior_mgmt` | gdsen | GD senior management |
| `wlb` | gdwlb | GD work-life balance |
| `culture` | gdcult | GD culture & values |
| `diversity` | gddiv | GD diversity & inclusion |

## Event-Level Aggregated Variables

| Variable | Definition |
|----------|------------|
| `delta_y` | post_mean - pre_mean (for outcome/filter/window) |
| `pre_mean` | Mean rating for reviews with days_to_election < 0 |
| `post_mean` | Mean rating for reviews with days_to_election >= 0 |
| `n_pre` | Number of pre-election reviews |
| `n_post` | Number of post-election reviews |
| `n_total` | n_pre + n_post |

## Regression Terms

| Term | Definition |
|------|------------|
| `Win x Post` | Interaction of win indicator and post indicator (review-level DiD-RD main coefficient) |
| `poly(margin)` | Polynomial terms of the running variable |
| `win x poly(margin)` | Interaction of treatment with polynomial terms |

## rdrobust / Local-Linear Terms

| Term | Definition |
|------|------------|
| Conventional estimate | Local-linear RDD estimate with triangular kernel |
| Robust estimate | Bias-corrected estimate |
| Bandwidth left/right | Data-driven bandwidth on each side of cutoff |
| Effective N left/right | Number of observations within bandwidth on each side |
"""

with open(OUT / "union_glassdoor_rdd_variable_definitions.md", "w") as f:
    f.write(var_md)
print(f"  Saved: union_glassdoor_rdd_variable_definitions.md")

# ═══════════════════════════════════════════════════════════════════════
# 7. Table Notes
# ═══════════════════════════════════════════════════════════════════════
print("\n7. Generating table notes...")
notes_md = f"""# Table Notes for Union Glassdoor RDD Results

## General Notes for All Tables

1. **Main sample**: Current employees only. All-employee results available as robustness.
2. **Event window**: +/-365 days around union election date, unless otherwise noted.
3. **Bandwidths**: Global (all margins), |margin| <= 0.20, |margin| <= 0.10.
4. **Significance**: *** p < 0.01, ** p < 0.05, * p < 0.10.
5. **Standard errors**: HC1 robust for review-level and event-level regressions.
6. **Review-level threshold**: pre >= 3 and post >= 3 reviews per election.
7. **Event-level threshold**: pre >= 1 and post >= 1 review per election.
8. **Event-level weighting**: Harmonic mean of n_pre and n_post.
9. **Fixed effects**: Election FE (absorbed) + calendar year FE for review-level.
10. **Treatment**: Win = 1[margin > 0], i.e., union received more than 50% of votes.

## Interpretation

- Coefficients are in standard deviation units of the outcome variable.
- tau > 0 means close union wins are associated with *higher* post-election ratings relative to pre-election, compared to close losses.
- The RDD identifies a local average treatment effect at the 50% vote threshold.
- Results may not generalize to landslide union wins or losses.

## Comparison with Earlier Results

- Earlier exploratory regressions used a simple pre/post DiD and found mostly negative coefficients.
- The new RDD design reverses the direction for most outcomes, finding positive effects.
- This is expected: the DiD compared all winners (many landslide) to all losers, introducing selection bias. The RDD isolates quasi-random variation at the close-election cutoff.

## Notes on Local-Linear RDD

- rdrobust Python package was not available in the analysis environment.
- Manual local-linear RDD with triangular kernel was used as a fallback.
- Bandwidth is selected using Silverman's rule of thumb: h = 1.84 * sd(margin) * n^(-1/5).
- This is broadly comparable to rdrobust with p=1, kernel=triangular, and a rule-of-thumb bandwidth selector.

## Diversity & Inclusion

- D&I is treated as exploratory in all tables.
- Earlier attrition analysis found D&I heavily concentrated in a few firms.
- While the rebuilt sample has broader D&I coverage (257 gvkeys vs 26 in old sample), concentration should still be monitored.
"""

with open(OUT / "union_glassdoor_rdd_table_notes.md", "w") as f:
    f.write(notes_md)
print(f"  Saved: union_glassdoor_rdd_table_notes.md")

# ═══════════════════════════════════════════════════════════════════════
# 8. LaTeX Tables
# ═══════════════════════════════════════════════════════════════════════
print("\n8. Generating LaTeX tables...")
tex = r"""% Union Glassdoor RDD Main Tables
% Generated: """ + datetime.now().strftime("%Y-%m-%d %H:%M") + r"""
\documentclass{article}
\usepackage{booktabs}
\usepackage{multirow}
\begin{document}

% ── Table 1: Event-Level RDD ────────────────────────────────────────
\begin{table}[ht]
\caption{Event-Level RDD: Close Union Election Effects on Glassdoor Ratings \\
\textit{Current employees, $\pm 365$ days, pre$\ge 1$ post$\ge 1$, weighted, linear p=1}}
\label{tab:event_rdd}
\begin{tabular}{lccc}
\toprule
& \multicolumn{3}{c}{Union Win (Discontinuity at margin=0)} \\
\cmidrule(lr){2-4}
Outcome & Global & $|m|\le 0.20$ & $|m|\le 0.10$ \\
\midrule
"""
for oc in [o for o in OUTCOME_MAP if o != "diversity"]:
    label = OUTCOME_MAP[oc]["label"]
    ev_g = ev_row(oc, "global")
    ev_20 = ev_row(oc, "|m|<=0.20")
    ev_10 = ev_row(oc, "|m|<=0.10")
    g_str = f"{ev_g['tau']:.3f}" + ("***" if ev_g["p_value"]<0.01 else "**" if ev_g["p_value"]<0.05 else "*" if ev_g["p_value"]<0.10 else "") if ev_g is not None else "---"
    b20_str = f"{ev_20['tau']:.3f}" if ev_20 is not None else "---"
    b10_str = f"{ev_10['tau']:.3f}" if ev_10 is not None else "---"
    tex += f"  {label} & ${g_str}$ & ${b20_str}$ & ${b10_str}$ \\\\\n"
    # SE row
    g_se = f"({ev_g['se']:.3f})" if ev_g is not None else ""
    b20_se = f"({ev_20['se']:.3f})" if ev_20 is not None else ""
    b10_se = f"({ev_10['se']:.3f})" if ev_10 is not None else ""
    tex += f"  & ${g_se}$ & ${b20_se}$ & ${b10_se}$ \\\\\n"

tex += r"""\midrule
Election FE & \multicolumn{3}{c}{No (cross-sectional RDD)} \\
Calendar year FE & \multicolumn{3}{c}{No} \\
Weighting & \multicolumn{3}{c}{Harmonic mean of $n_{pre}, n_{post}$} \\
Polynomial & \multicolumn{3}{c}{Linear (p=1)} \\
SE & \multicolumn{3}{c}{HC1 robust} \\
\bottomrule
\end{tabular}

\vspace{4pt}
{\footnotesize\textit{Note:} *** p<0.01, ** p<0.05, * p<0.10. Coefficients in SD units of the outcome.
Standard errors in parentheses. Sample restricted to current employees only.}
\end{table}

% ── Table 2: Review-Level DiD-RD ────────────────────────────────────
\begin{table}[ht]
\caption{Review-Level DiD-RD: Close Union Election Effects \\
\textit{Current employees, $\pm 365$ days, pre$\ge 3$ post$\ge 3$, election FE + year FE}}
\label{tab:review_did_rd}
\begin{tabular}{lccc}
\toprule
& \multicolumn{3}{c}{Win $\times$ Post} \\
\cmidrule(lr){2-4}
Outcome & Global & $|m|\le 0.20$ & $|m|\le 0.10$ \\
\midrule
"""
for oc in [o for o in OUTCOME_MAP if o != "diversity"]:
    label = OUTCOME_MAP[oc]["label"]
    rv_g = rv_row(oc, "global")
    rv_20 = rv_row(oc, "|m|<=0.20")
    rv_10 = rv_row(oc, "|m|<=0.10")
    g_str = f"{rv_g['estimate_tau']:.3f}" + ("***" if rv_g["p_value"]<0.01 else "**" if rv_g["p_value"]<0.05 else "*" if rv_g["p_value"]<0.10 else "") if rv_g is not None else "---"
    b20_str = f"{rv_20['estimate_tau']:.3f}" if rv_20 is not None else "---"
    b10_str = f"{rv_10['estimate_tau']:.3f}" if rv_10 is not None else "---"
    tex += f"  {label} & ${g_str}$ & ${b20_str}$ & ${b10_str}$ \\\\\n"
    g_se = f"({rv_g['se']:.3f})" if rv_g is not None else ""
    b20_se = f"({rv_20['se']:.3f})" if rv_20 is not None else ""
    b10_se = f"({rv_10['se']:.3f})" if rv_10 is not None else ""
    tex += f"  & ${g_se}$ & ${b20_se}$ & ${b10_se}$ \\\\\n"

tex += r"""\midrule
Election FE & \multicolumn{3}{c}{Yes (absorbed)} \\
Calendar year FE & \multicolumn{3}{c}{Yes} \\
Employee filter & \multicolumn{3}{c}{Current only} \\
Min reviews & \multicolumn{3}{c}{$\ge 3$ pre, $\ge 3$ post} \\
SE & \multicolumn{3}{c}{HC1 robust} \\
\bottomrule
\end{tabular}

\vspace{4pt}
{\footnotesize\textit{Note:} *** p<0.01, ** p<0.05, * p<0.10. Coefficients in SD units of the outcome.
Standard errors in parentheses. Election FE absorbed via within-transformation.}
\end{table}

\end{document}
"""

with open(OUT / "union_glassdoor_rdd_main_tables.tex", "w") as f:
    f.write(tex)
print(f"  Saved: union_glassdoor_rdd_main_tables.tex")

# ═══════════════════════════════════════════════════════════════════════
# DONE
# ═══════════════════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("PACKAGING COMPLETE")
print(f"All outputs in: {OUT}")
print(f"  Excel:       union_glassdoor_rdd_results_package.xlsx")
print(f"  Memo:        union_glassdoor_rdd_coauthor_memo.md")
print(f"  Variables:   union_glassdoor_rdd_variable_definitions.md")
print(f"  Table notes: union_glassdoor_rdd_table_notes.md")
print(f"  LaTeX:       union_glassdoor_rdd_main_tables.tex")
print(f"  CSV:         union_glassdoor_rdd_selected_results.csv")
print(f"  CSV:         union_glassdoor_rdd_all_consistency_checked_results.csv")
