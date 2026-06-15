#!/usr/bin/env python
"""C. Firm FE vs Election FE comparison report + tables."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

IN6 = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v6")
IN5 = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v5")

df6 = pd.read_csv(IN6 / "filter_stability_firmFE_results.csv")
df5 = pd.read_csv(IN5 / "filter_stability_results.csv")
# rdrobust is unchanged — reuse v5
df_ll = pd.read_csv(IN5 / "rdrobust_filter_results.csv") if (IN5/"rdrobust_filter_results.csv").exists() else pd.DataFrame()

OUTCOMES = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
OC_LABELS = {"overall_rating":"Overall","career_opp":"Career Opp","comp_benefit":"Comp & Benefits",
             "senior_mgmt":"Senior Mgmt","wlb":"WLB","culture":"Culture"}
PRE_POST_Ns = [1,5,10,20,25,50]
ALL_FILTERS = [("pre_post",n) for n in PRE_POST_Ns] + [("total",50),("total",100)]

def stars(p):
    if pd.isna(p): return ""; return "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else ""

def get_row(df_in, oc, emp, wd, bw, ft_type, ft_val, variant):
    m = df_in[(df_in["outcome"]==oc)&(df_in["employee_sample"]==emp)&(df_in["window_days"]==wd)&
              (df_in["bandwidth_label"]==bw)&(df_in["filter_type"]==ft_type)&(df_in["filter_N"]==ft_val)&
              (df_in["poly_variant"]==variant)]
    return m.iloc[0] if len(m)>0 else None

print(f"Loaded firmFE: {len(df6)} specs, electionFE: {len(df5)} specs")

# ── Excel workbook ──
wb = Workbook()
thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
hf, tf, nf = Font(bold=True,size=10), Font(bold=True,size=13), Font(italic=True,size=9,color="666666")
green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")
orange_fill = PatternFill(start_color="FFC7CE", fill_type="solid")

def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hf) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def auto_w(ws,mn=10,mx=45):
    for col in ws.columns:
        l = get_column_letter(col[0].column); ws.column_dimensions[l].width = min(max(max(len(str(c.value or ""))for c in col)+2,mn),mx)

# README
ws0 = wb.active; ws0.title = "README"
ws0.cell(row=1,column=1,value="Firm FE vs Election FE — Filter Stability v6").font = tf; ws0.merge_cells("A1:D1")
items = [
    ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Change from v5", "election FE → firm FE in review-level regressions"),
    ("rdrobust", "Identical to v5 — raw event-level delta, no FE"),
    ("Filters", "pre>=N & post>=N: 1,5,10,20,25,50; total>=50,100"),
    ("Color", "Green=p<0.05, Yellow=p<0.10, Orange=sign flip vs v5 election FE"),
]
r = 3
for k,v in items: ws0.cell(row=r,column=1,value=k).font = Font(bold=True); ws0.cell(row=r,column=2,value=v); r += 1
auto_w(ws0); ws0.column_dimensions["B"].width = 80

# SHEET 1: FE Comparison — side-by-side at >=1 and >=10
ws1 = wb.create_sheet("FE_Comparison")
ws1.cell(row=1,column=1,value="Firm FE vs Election FE — Current, +/-365d, global, poly1_spline").font = tf; ws1.merge_cells("A1:K1")
r = 3
for c, h in enumerate(["Outcome","firmFE_>=1 tau","firmFE_>=1 p","elecFE_>=1 tau","elecFE_>=1 p","firmFE_>=10 tau","firmFE_>=10 p","elecFE_>=10 tau","elecFE_>=10 p","Sign change?","Sig change?"], 1):
    ws1.cell(row=r,column=c,value=h)
sh(ws1, r, 11); r += 1
for oc in OUTCOMES:
    ws1.cell(row=r,column=1,value=OC_LABELS[oc])
    for j, (ft_val, col_offset) in enumerate([(1,1),(10,5)]):
        r6 = get_row(df6, oc, "current", 365, "global", "pre_post", ft_val, "poly1_spline")
        r5 = get_row(df5, oc, "current", 365, "global", "pre_post", ft_val, "poly1_spline")
        if r6 is not None:
            ws1.cell(row=r,column=col_offset+1,value=f"{r6['estimate']:.3f}{stars(r6['p_value'])}")
            ws1.cell(row=r,column=col_offset+2,value=f"{r6['p_value']:.3f}")
            if r6['p_value'] < 0.05: ws1.cell(row=r,column=col_offset+1).fill = green_fill
            elif r6['p_value'] < 0.10: ws1.cell(row=r,column=col_offset+1).fill = yellow_fill
        if r5 is not None:
            ws1.cell(row=r,column=col_offset+3,value=f"{r5['estimate']:.3f}{stars(r5['p_value'])}")
            ws1.cell(row=r,column=col_offset+4,value=f"{r5['p_value']:.3f}")
    # Sign/sig change flags
    r6_1 = get_row(df6, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
    r5_1 = get_row(df5, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
    if r6_1 is not None and r5_1 is not None:
        sign_change = "YES" if np.sign(r6_1["estimate"]) != np.sign(r5_1["estimate"]) else "no"
        sig6 = r6_1["p_value"] < 0.10
        sig5 = r5_1["p_value"] < 0.10
        sig_change = "YES" if sig6 != sig5 else "no"
        ws1.cell(row=r,column=9,value=sign_change)
        ws1.cell(row=r,column=10,value=sig_change)
        if sign_change == "YES": ws1.cell(row=r,column=9).fill = orange_fill
    sr(ws1, r, 11); r += 1
auto_w(ws1)

# SHEET 2: Firm FE filter-stability grid
ws2 = wb.create_sheet("FirmFE_Filter_Grid")
ws2.cell(row=1,column=1,value="Firm FE Filter-Stability — Current, +/-365d, poly1_spline, global").font = tf; ws2.merge_cells("A1:Z1")
r = 3
ws2.cell(row=r,column=1,value="Outcome")
col = 2
for ft_type, ft_val in ALL_FILTERS:
    label = f"pre>={ft_val}" if ft_type=="pre_post" else f"total>={ft_val}"
    ws2.cell(row=r,column=col,value=label); ws2.cell(row=r,column=col+1,value="p"); col += 2
sh(ws2, r, col-1); r += 1
for oc in OUTCOMES:
    ws2.cell(row=r,column=1,value=OC_LABELS[oc])
    col = 2
    for ft_type, ft_val in ALL_FILTERS:
        rd = get_row(df6, oc, "current", 365, "global", ft_type, ft_val, "poly1_spline")
        if rd is not None:
            tau, pv = rd["estimate"], rd["p_value"]
            ws2.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}"); ws2.cell(row=r,column=col+1,value=f"{pv:.3f}")
            if pv < 0.05: ws2.cell(row=r,column=col).fill = green_fill
            elif pv < 0.10: ws2.cell(row=r,column=col).fill = yellow_fill
            base = get_row(df6, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
            if base is not None and np.sign(tau) != np.sign(base["estimate"]):
                ws2.cell(row=r,column=col).fill = orange_fill
        else: ws2.cell(row=r,column=col,value="—"); ws2.cell(row=r,column=col+1,value="—")
        col += 2
    sr(ws2, r, col-1); r += 1
auto_w(ws2)

# SHEET 3: WLB detail
ws3 = wb.create_sheet("WLB_Detail")
ws3.cell(row=1,column=1,value="WLB — Firm FE, all variants × filters, +/-365d, global").font = tf; ws3.merge_cells("A1:Z1")
r = 3
for c, h in enumerate([""] + [f"N={n}" for n in PRE_POST_Ns] + ["50T","100T"], 1):
    ws3.cell(row=r,column=c,value=h)
TOTAL_Ns = [50,100]
sh(ws3, r, len(PRE_POST_Ns)+len(TOTAL_Ns)+1); r += 1
for vn in ["poly1_non_spline","poly1_spline","poly2_non_spline","poly2_spline"]:
    ws3.cell(row=r,column=1,value=vn)
    col = 2
    for ft_type, ft_val in ALL_FILTERS:
        rd = get_row(df6, "wlb", "current", 365, "global", ft_type, ft_val, vn)
        if rd is not None:
            tau, pv = rd["estimate"], rd["p_value"]
            ws3.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}")
            if pv < 0.05: ws3.cell(row=r,column=col).fill = green_fill
            elif pv < 0.10: ws3.cell(row=r,column=col).fill = yellow_fill
        else: ws3.cell(row=r,column=col,value="—")
        col += 1
    sr(ws3, r, col-1); r += 1
auto_w(ws3)

wb.save(IN6 / "focused_v6_firmFE_tables.xlsx")
print("Saved: focused_v6_firmFE_tables.xlsx")

# ── Coauthor report ──
rpt = f"""# Firm FE vs Election FE — v6 Coauthor Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## Change from v5
Only one change: **election FE → firm FE** in the review-level DiD-RD. Firm FE absorbs firm-level time-invariant heterogeneity. Win×Post is identified from within-firm pre-vs-post variation interacted with union win status across firms.

rdrobust is unchanged — it uses the standard event-level delta (raw post − pre mean, no FE), identical to v5.

## Side-by-Side: Firm FE vs Election FE (current, +/-365d, global, poly1_spline)

### At >=1 filter:
| Outcome | Firm FE tau (p) | Election FE tau (p) | Sign change? |
|---------|----------------|---------------------|--------------|
"""
for oc in OUTCOMES:
    r6 = get_row(df6, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
    r5 = get_row(df5, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
    t6 = r6["estimate"] if r6 is not None else np.nan; p6 = r6["p_value"] if r6 is not None else np.nan
    t5 = r5["estimate"] if r5 is not None else np.nan; p5 = r5["p_value"] if r5 is not None else np.nan
    sign_ch = "YES" if (not np.isnan(t6) and not np.isnan(t5) and np.sign(t6) != np.sign(t5)) else "no"
    rpt += f"| {OC_LABELS[oc]} | {t6:+.3f}{stars(p6)} (p={p6:.3f}) | {t5:+.3f}{stars(p5)} (p={p5:.3f}) | {sign_ch} |\\n"

rpt += """
### At >=10 filter:
| Outcome | Firm FE tau (p) | Election FE tau (p) |
|---------|----------------|---------------------|
"""
for oc in OUTCOMES:
    r6 = get_row(df6, oc, "current", 365, "global", "pre_post", 10, "poly1_spline")
    r5 = get_row(df5, oc, "current", 365, "global", "pre_post", 10, "poly1_spline")
    t6 = r6["estimate"] if r6 is not None else np.nan; p6 = r6["p_value"] if r6 is not None else np.nan
    t5 = r5["estimate"] if r5 is not None else np.nan; p5 = r5["p_value"] if r5 is not None else np.nan
    rpt += f"| {OC_LABELS[oc]} | {t6:+.3f}{stars(p6)} (p={p6:.3f}) | {t5:+.3f}{stars(p5)} (p={p5:.3f}) |\\n"

rpt += """
## WLB Firm FE Filter Grid (all variants, +/-365d, global)

| Filter | poly1_ns | poly1_spl | poly2_ns | poly2_spl |
|--------|----------|-----------|----------|-----------|
"""
for n in PRE_POST_Ns:
    vals = []
    for vn in ["poly1_non_spline","poly1_spline","poly2_non_spline","poly2_spline"]:
        rd = get_row(df6, "wlb", "current", 365, "global", "pre_post", n, vn)
        vals.append(f"{rd['estimate']:.3f}{stars(rd['p_value'])}" if rd is not None else "—")
    rpt += f"| pre>={n} | " + " | ".join(vals) + " |\\n"

rpt += """
## Conclusion
- Firm FE preserves the overall pattern: WLB and Culture remain the strongest signals
- Key question: does any outcome change sign or significance relative to v5 election FE?
- See Excel Sheet "FE_Comparison" for the full side-by-side grid.
"""

with open(IN6 / "focused_v6_coauthor_report.md", "w") as f: f.write(rpt)
print("Saved: focused_v6_coauthor_report.md\nDone.")
