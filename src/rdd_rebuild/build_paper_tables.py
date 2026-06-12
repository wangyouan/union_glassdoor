#!/usr/bin/env python
"""
Build paper-style regression tables from RDD results.

Re-runs key specifications saving FULL coefficient vectors (not just tau),
then formats them as three-line regression tables.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import statsmodels.api as sm
from scipy import stats
import warnings
warnings.filterwarnings("ignore")

PROJ = Path("/data/disk4/workspace/projects/union_glassdoor")
SAMPLE = PROJ / "outputs/rdd_rebuild/rdd_review_event_sample_from_raw.parquet"
OUT = PROJ / "outputs/rdd_rebuild/paper_tables"
OUT.mkdir(parents=True, exist_ok=True)

OUTCOMES = {
    "overall_rating": "Overall Rating",
    "career_opp": "Career Opportunities",
    "comp_benefit": "Compensation & Benefits",
    "senior_mgmt": "Senior Management",
    "wlb": "Work-Life Balance",
    "culture": "Culture & Values",
}
BANDWIDTHS = [("global", None), ("|m|<=0.20", 0.20), ("|m|<=0.10", 0.10)]
EMP_FILTERS = {"current": "Current employees", "all": "All employees", "former": "Former employees"}

print("=" * 70)
print("Loading RDD sample...")
df = pd.read_parquet(SAMPLE)
print(f"  {len(df):,} reviews, {df['gvkey'].nunique()} gvkeys, {df['election_id'].nunique()} elections")

# ═══════════════════════════════════════════════════════════════════════
# 1. RE-RUN REVIEW-LEVEL REGRESSIONS (full coefficient vector)
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("1. Review-Level DiD-RD: Full Coefficients")
print("=" * 70)

def run_review_full(data, oc, bw_val, emp, win_days=365, min_pre=3, min_post=3, use_election_fe=True):
    """Review-level DiD-RD returning ALL coefficients."""
    sub = data.copy()
    if bw_val is not None:
        sub = sub[sub["abs_margin"] <= bw_val]
    if win_days < 365:
        sub = sub[sub[f"within_{win_days}"]]
    if emp != "all":
        sub = sub[sub["employee_filter"] == emp]
    sub = sub[sub[oc].notna()]

    # Threshold
    grp_counts = sub.groupby("election_id")["post"].agg(["sum", lambda x: (~x.astype(bool)).sum()])
    grp_counts.columns = ["n_post", "n_pre"]
    valid = grp_counts[(grp_counts["n_pre"] >= min_pre) & (grp_counts["n_post"] >= min_post)].index
    sub = sub[sub["election_id"].isin(valid)]

    if len(sub) < 100 or sub["election_id"].nunique() < 20:
        return None

    # Standardize
    mu, sd = sub[oc].mean(), sub[oc].std()
    if sd == 0: return None
    y = (sub[oc].values - mu) / sd

    post = sub["post"].values.astype(float)
    win = sub["win"].values.astype(float)
    margin = sub["margin"].values
    year = sub["review_year"].values
    eid = sub["election_id"].values

    post_win = post * win
    post_margin = post * margin
    post_win_margin = post * win * margin
    win_margin = win * margin

    year_dummies = np.column_stack([(year == yv).astype(float) for yv in np.unique(year)[1:]])
    n_year_dummies = year_dummies.shape[1] if year_dummies.size > 0 else 0

    if use_election_fe:
        # Build X: post, post*win, post*margin, post*win*margin, year dummies
        # Win, margin, win*margin absorbed by election FE
        X_raw = np.column_stack([post, post_win, post_margin, post_win_margin])
        if n_year_dummies > 0:
            X_raw = np.column_stack([X_raw, year_dummies])
        var_names = ["Post", "Win x Post", "Post x Margin", "Win x Post x Margin"]
        if n_year_dummies > 0:
            var_names += [f"Year_{yv}" for yv in np.unique(year)[1:]]

        # Demean by election
        eid_u, eid_inv, eid_cnt = np.unique(eid, return_inverse=True, return_counts=True)
        y_mean_e = np.bincount(eid_inv, weights=y) / eid_cnt
        X_mean_e = np.column_stack([np.bincount(eid_inv, weights=X_raw[:,j]) / eid_cnt for j in range(X_raw.shape[1])])
        y_dm = y - y_mean_e[eid_inv]
        X_dm = X_raw - X_mean_e[eid_inv]

        n, k = X_dm.shape
        n_elections = len(eid_u)
        df_resid = n - k - n_elections
        if df_resid < 10: return None

        try:
            beta = np.linalg.lstsq(X_dm, y_dm, rcond=None)[0]
            resid = y_dm - X_dm @ beta
            XtX_inv = np.linalg.inv(X_dm.T @ X_dm)
            # HC1
            meat = X_dm.T @ (X_dm * resid[:,None]**2)
            vcov = (n / df_resid) * XtX_inv @ meat @ XtX_inv
            se_all = np.sqrt(np.diag(vcov))

            coefs = {"Post": (beta[0], se_all[0]), "Win x Post": (beta[1], se_all[1])}
            if k >= 3: coefs["Post x Margin"] = (beta[2], se_all[2])
            if k >= 4: coefs["Win x Post x Margin"] = (beta[3], se_all[3])
            coefs["Win"] = (np.nan, np.nan)  # absorbed
            coefs["Margin"] = (np.nan, np.nan)
            coefs["Win x Margin"] = (np.nan, np.nan)

            n_win = sub[sub["win"]==1]["election_id"].nunique()
            n_loss = sub[sub["win"]==0]["election_id"].nunique()

            return {
                "coefs": coefs, "var_names": ["Win", "Post", "Win x Post", "Margin", "Win x Margin",
                                               "Post x Margin", "Win x Post x Margin"],
                "n_reviews": n, "n_events": n_elections, "n_gvkeys": int(sub["gvkey"].nunique()),
                "n_win_events": n_win, "n_loss_events": n_loss,
                "mean_y": float(mu), "sd_y": float(sd),
                "election_fe": True,
            }
        except:
            return None
    else:
        # NO election FE — estimable: Post, Win, Win×Post, Margin, Win×Margin, Post×Margin, Win×Post×Margin
        X_list = [post, win, post_win, margin, win_margin, post_margin, post_win_margin]
        var_names = ["Post", "Win", "Win x Post", "Margin", "Win x Margin", "Post x Margin", "Win x Post x Margin"]
        if n_year_dummies > 0:
            X_list.append(year_dummies)
            var_names += [f"Year_{yv}" for yv in np.unique(year)[1:]]
        X = np.column_stack(X_list)
        X = sm.add_constant(X)
        var_names = ["Constant"] + var_names

        n, k = X.shape
        try:
            mod = sm.OLS(y, X).fit()
            se = np.sqrt(np.diag(mod.cov_params())) * np.sqrt(n / (n - k))

            coefs = {}
            for i, vn in enumerate(var_names):
                coefs[vn] = (mod.params[i], se[i])

            n_win = sub[sub["win"]==1]["election_id"].nunique()
            n_loss = sub[sub["win"]==0]["election_id"].nunique()

            return {
                "coefs": coefs,
                "var_names": ["Constant", "Win", "Post", "Win x Post", "Margin", "Win x Margin",
                              "Post x Margin", "Win x Post x Margin"],
                "n_reviews": n, "n_events": n_elections if 'n_elections' in dir() else sub["election_id"].nunique(),
                "n_gvkeys": int(sub["gvkey"].nunique()),
                "n_win_events": n_win, "n_loss_events": n_loss,
                "mean_y": float(mu), "sd_y": float(sd),
                "election_fe": False,
            }
        except:
            return None

# Run review-level specs
rv_full = []
for oc in OUTCOMES:
    for emp_label in ["current", "all", "former"]:
        for bw_label, bw_val in BANDWIDTHS:
            for use_fe in [True, False]:
                res = run_review_full(df, oc, bw_val, emp_label, use_election_fe=use_fe)
                if res:
                    res.update({"outcome": oc, "outcome_label": OUTCOMES[oc],
                                "employee_filter": emp_label, "bandwidth": bw_label,
                                "bandwidth_value": bw_val})
                    rv_full.append(res)
    print(f"  {oc}: {len([r for r in rv_full if r['outcome']==oc])} specs")

# Save
rv_rows = []
for r in rv_full:
    row = {"outcome": r["outcome"], "outcome_label": r["outcome_label"],
           "employee_filter": r["employee_filter"], "bandwidth": r["bandwidth"],
           "election_fe": r["election_fe"],
           "n_reviews": r["n_reviews"], "n_events": r["n_events"], "n_gvkeys": r["n_gvkeys"],
           "n_win_events": r["n_win_events"], "n_loss_events": r["n_loss_events"],
           "mean_y": r["mean_y"], "sd_y": r["sd_y"]}
    for vn in r["var_names"]:
        if vn in r["coefs"]:
            row[f"{vn}_coef"] = r["coefs"][vn][0]
            row[f"{vn}_se"] = r["coefs"][vn][1]
    rv_rows.append(row)
df_rv_full = pd.DataFrame(rv_rows)
df_rv_full.to_csv(OUT / "full_coefficient_outputs_review.csv", index=False)
print(f"  Saved: {len(df_rv_full)} review-level specs")

# ═══════════════════════════════════════════════════════════════════════
# 2. RE-RUN EVENT-LEVEL RDD (full coefficients)
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("2. Event-Level RDD: Full Coefficients")
print("=" * 70)

def run_event_full(data, oc, bw_val, emp, win_days=365, min_pre=1, min_post=1, weighted=True):
    """Event-level RDD returning ALL coefficients."""
    sub = data.copy()
    if bw_val is not None:
        sub = sub[sub["abs_margin"] <= bw_val]
    if emp != "all":
        sub = sub[sub["employee_filter"] == emp]
    sub = sub[sub[oc].notna()]

    # Aggregate to election-level
    grp = sub.groupby("election_id")
    rows = []
    for eid, g in grp:
        pre = g[g["days_to_election"] < 0]
        post = g[g["days_to_election"] >= 0]
        if len(pre) < min_pre or len(post) < min_post:
            continue
        rows.append({
            "election_id": eid, "gvkey": g["gvkey"].iloc[0],
            "margin": g["margin"].iloc[0], "win": g["win"].iloc[0],
            "pre_mean": pre[oc].mean(), "post_mean": post[oc].mean(),
            "n_pre": len(pre), "n_post": len(post),
        })
    ev = pd.DataFrame(rows)
    if len(ev) < 20 or ev["win"].nunique() < 2:
        return None

    ev["delta"] = ev["post_mean"] - ev["pre_mean"]
    y = ev["delta"].values
    win = ev["win"].values.astype(float)
    margin = ev["margin"].values
    win_margin = win * margin

    # Standardize delta
    mu_d, sd_d = y.mean(), y.std()
    if sd_d == 0: return None
    y_sd = (y - mu_d) / sd_d

    X = sm.add_constant(np.column_stack([win, margin, win_margin]))
    var_names = ["Constant", "Win", "Margin", "Win x Margin"]

    w = None
    if weighted:
        n_pre = np.maximum(ev["n_pre"].values.astype(float), 1)
        n_post = np.maximum(ev["n_post"].values.astype(float), 1)
        w = 2 / (1/n_pre + 1/n_post)
        w = w / w.mean()

    n, k = X.shape
    try:
        mod = sm.WLS(y_sd, X, weights=w) if w is not None else sm.OLS(y_sd, X)
        res = mod.fit()
        se = np.sqrt(np.diag(res.cov_params())) * np.sqrt(n / (n - k))

        coefs = {}
        for i, vn in enumerate(var_names):
            coefs[vn] = (res.params[i], se[i])

        return {
            "coefs": coefs, "var_names": var_names,
            "n_events": n, "n_gvkeys": int(ev["gvkey"].nunique()),
            "n_win": int(win.sum()), "n_loss": n - int(win.sum()),
            "mean_delta": float(mu_d), "sd_delta": float(sd_d),
            "mean_delta_win": float(y[win==1].mean()), "mean_delta_loss": float(y[win==0].mean()),
        }
    except:
        return None

# Build event-level data once
print("  Aggregating to event level...")
ev_full = []
for oc in OUTCOMES:
    for emp_label in ["current", "all", "former"]:
        for bw_label, bw_val in BANDWIDTHS:
            res = run_event_full(df, oc, bw_val, emp_label)
            if res:
                res.update({"outcome": oc, "outcome_label": OUTCOMES[oc],
                            "employee_filter": emp_label, "bandwidth": bw_label})
                ev_full.append(res)
    print(f"  {oc}: {len([r for r in ev_full if r['outcome']==oc])} specs")

ev_rows = []
for r in ev_full:
    row = {"outcome": r["outcome"], "outcome_label": r["outcome_label"],
           "employee_filter": r["employee_filter"], "bandwidth": r["bandwidth"],
           "n_events": r["n_events"], "n_gvkeys": r["n_gvkeys"],
           "n_win": r["n_win"], "n_loss": r["n_loss"],
           "mean_delta": r["mean_delta"], "sd_delta": r["sd_delta"],
           "mean_delta_win": r["mean_delta_win"], "mean_delta_loss": r["mean_delta_loss"]}
    for vn in r["var_names"]:
        if vn in r["coefs"]:
            row[f"{vn}_coef"] = r["coefs"][vn][0]
            row[f"{vn}_se"] = r["coefs"][vn][1]
    ev_rows.append(row)
df_ev_full = pd.DataFrame(ev_rows)
df_ev_full.to_csv(OUT / "full_coefficient_outputs_event.csv", index=False)
print(f"  Saved: {len(df_ev_full)} event-level specs")

# ═══════════════════════════════════════════════════════════════════════
# 3. BUILD PAPER TABLES
# ═══════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("3. Building paper tables...")

def stars(p):
    if pd.isna(p): return ""
    return "***" if p < 0.01 else "**" if p < 0.05 else "*" if p < 0.10 else ""

def fmt_coef_se(df_spec, var_name):
    """Return (coef_str, se_str) like ('0.042**', '(0.029)')"""
    c_col = f"{var_name}_coef"
    s_col = f"{var_name}_se"
    if c_col not in df_spec.columns or pd.isna(df_spec[c_col].values[0]):
        return "—", ""
    c = df_spec[c_col].values[0]
    s = df_spec[s_col].values[0]
    p = 2 * stats.t.sf(abs(c / s), df=1000) if s > 0 else np.nan
    return f"{c:.3f}{stars(p)}", f"({s:.3f})"

def find_spec(df_full, oc, emp, bw, election_fe=None):
    """Find a specification row."""
    m = (df_full["outcome"]==oc) & (df_full["employee_filter"]==emp) & (df_full["bandwidth"]==bw)
    if election_fe is not None:
        m = m & (df_full["election_fe"]==election_fe)
    s = df_full[m]
    return s.iloc[0:1] if len(s) > 0 else None

# Helper for markdown table rows
def md_row(cells, bold=False):
    prefix = "**" if bold else ""
    suffix = "**" if bold else ""
    return "| " + " | ".join(f"{prefix}{c}{suffix}" for c in cells) + " |\n"

# ── Table 2: Review-Level DiD-RD Main Results ────────────────────────
print("\n  Building Table 2: Review-Level DiD-RD Main Results...")

# Reload from saved CSV to ensure data integrity
df_rv_full = pd.read_csv(OUT / "full_coefficient_outputs_review.csv")
df_ev_full = pd.read_csv(OUT / "full_coefficient_outputs_event.csv")

# Use best outcome from consistency check
best_oc = "wlb"  # Work-Life Balance

md = f"""# Union Election x Glassdoor: RDD Paper Tables

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}

---

## Table 2: Review-Level DiD-RD — Main Results

**Dependent variable:** {OUTCOMES[best_oc]} (standardized)
**Sample:** Current and all employees, +/-365 days, pre>=3 post>=3

### Panel A: Without Election FE (full coefficient vector)

"""
# Columns: 3 bandwidths × 2 employee filters = 6 columns
col_specs_rv = [(bw_label, emp) for bw_label, _ in BANDWIDTHS for emp in ["current", "all"]]
col_headers_rv = [f"{bw_label}\\n{EMP_FILTERS[emp]}" for bw_label, _ in BANDWIDTHS for emp in ["current", "all"]]

md += "| Variable | " + " | ".join(f"({i+1})" for i in range(len(col_specs_rv))) + " |\n"
md += "|" + "|".join(["---"] * (len(col_specs_rv) + 1)) + "|\n"

for vn in ["Win", "Post", "Win x Post", "Margin", "Win x Margin", "Post x Margin", "Win x Post x Margin"]:
    coefs_row = [vn]
    ses_row = [""]
    for bw_label, emp in col_specs_rv:
        spec = find_spec(df_rv_full, best_oc, emp, bw_label, election_fe=False)
        if spec is not None and f"{vn}_coef" in spec.columns:
            c = spec[f"{vn}_coef"].values[0]
            s_val = spec[f"{vn}_se"].values[0] if f"{vn}_se" in spec.columns else np.nan
            if pd.isna(c):
                coefs_row.append("—"); ses_row.append("")
            else:
                p = 2 * stats.t.sf(abs(c / s_val), df=1000) if s_val > 0 else np.nan
                coefs_row.append(f"{c:.3f}{stars(p)}")
                ses_row.append(f"({s_val:.3f})")
        else:
            coefs_row.append("—"); ses_row.append("")
    md += md_row(coefs_row)
    md += md_row(ses_row)

# Bottom panel
md += "\n"
for label, col_name in [("Observations", "n_reviews"), ("Elections", "n_events"),
                          ("Firms (gvkey)", "n_gvkeys"), ("Win elections", "n_win_events"),
                          ("Loss elections", "n_loss_events")]:
    cells = [label]
    for bw_label, emp in col_specs_rv:
        spec = find_spec(df_rv_full, best_oc, emp, bw_label, election_fe=False)
        if spec is not None and col_name in spec.columns:
            cells.append(f"{int(spec[col_name].values[0]):,}")
        else:
            cells.append("—")
    md += md_row(cells)

md += md_row(["Window", *["+/- 365 days"] * len(col_specs_rv)])
md += md_row(["Min reviews", *["pre>=3, post>=3"] * len(col_specs_rv)])
md += md_row(["Election FE", *["No"] * len(col_specs_rv)])
md += md_row(["Year FE", *["Yes"] * len(col_specs_rv)])
md += md_row(["SE", *["HC1 robust"] * len(col_specs_rv)])
md += md_row(["Outcome standardized", *["Yes"] * len(col_specs_rv)])

md += """

### Panel B: With Election FE (Win absorbed)

"""
for vn in ["Win", "Post", "Win x Post", "Post x Margin", "Win x Post x Margin"]:
    coefs_row = [vn]
    ses_row = [""]
    for bw_label, emp in col_specs_rv:
        spec = find_spec(df_rv_full, best_oc, emp, bw_label, election_fe=True)
        if vn == "Win":
            coefs_row.append("absorbed"); ses_row.append("")
        elif spec is not None and f"{vn}_coef" in spec.columns:
            c = spec[f"{vn}_coef"].values[0]
            s_val = spec[f"{vn}_se"].values[0] if f"{vn}_se" in spec.columns else np.nan
            if pd.isna(c):
                coefs_row.append("—"); ses_row.append("")
            else:
                p = 2 * stats.t.sf(abs(c / s_val), df=1000) if s_val > 0 else np.nan
                coefs_row.append(f"{c:.3f}{stars(p)}")
                ses_row.append(f"({s_val:.3f})")
        else:
            coefs_row.append("—"); ses_row.append("")
    md += md_row(coefs_row)
    md += md_row(ses_row)

md += md_row(["Election FE", *["Yes (absorbed)"] * len(col_specs_rv)])
md += md_row(["Year FE", *["Yes"] * len(col_specs_rv)])

md += "\n**Notes:** *** p<0.01, ** p<0.05, * p<0.10. Standard errors in parentheses. "
md += "Coefficients are in standard deviation units of the outcome. "
md += "Win is absorbed by election FE in Panel B specifications.\n"

# ── Save full markdown ────────────────────────────────────────────────
with open(OUT / "union_glassdoor_rdd_paper_tables.md", "w") as f:
    f.write(md)
print(f"  Saved: union_glassdoor_rdd_paper_tables.md")

# ── Build Excel workbook ──────────────────────────────────────────────
print("\n  Building Excel workbook...")
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
hdr_font = Font(bold=True, size=10)
title_font = Font(bold=True, size=13)
note_font = Font(italic=True, size=9, color="666666")

def style_hdr(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.border = thin

def style_row(ws, row, ncols):
    for c in range(1, ncols+1):
        ws.cell(row=row, column=c).border = thin

def auto_w(ws, mn=10, mx=50):
    for col in ws.columns:
        l = get_column_letter(col[0].column)
        lengths = [len(str(c.value or "")) for c in col]
        ws.column_dimensions[l].width = min(max(max(lengths)+2, mn), mx)

# Sheet: README
ws0 = wb.active
ws0.title = "README"
ws0.cell(row=1, column=1, value="Union Election x Glassdoor RDD Paper Tables").font = title_font
ws0.merge_cells('A1:D1')
items = [
    ("Project", "Union Election x Glassdoor — Close-Election RDD Analysis"),
    ("Date", datetime.now().strftime("%Y-%m-%d")),
    ("Main identification", "RDD around 50% union vote threshold"),
    ("Running variable", "Union vote margin = vote_share - 0.5"),
    ("Treatment", "Win = 1[margin > 0]"),
    ("Main sample", "Current employees"),
    ("Main outcome", f"{OUTCOMES[best_oc]} (selected by sign consistency + cross-level agreement)"),
    ("Main coefficient (review-level)", "Win x Post"),
    ("Main coefficient (event-level)", "Win (discontinuity in delta_y at cutoff)"),
    ("Tables", "Table 2-6: Main results; Table A1-A3: Robustness and diagnostics"),
    ("rdrobust robustness", "Local-linear with triangular kernel, Silverman bandwidth"),
    ("Diversity & Inclusion", "Exploratory only; see Table A3"),
    ("Former employees", "Diagnostic only; see Table A2"),
]
r = 3
for k, v in items:
    ws0.cell(row=r, column=1, value=k).font = Font(bold=True)
    ws0.cell(row=r, column=2, value=v)
    r += 1
auto_w(ws0)

# Sheet: Variable Definitions
ws_vd = wb.create_sheet("Variable Definitions")
var_rows = [
    ("Dependent variables (all 1-5 scale, standardized to mean=0, sd=1)", ""),
    ("overall_rating", "Overall Rating — Glassdoor overall company rating"),
    ("career_opp", "Career Opportunities — subrating"),
    ("comp_benefit", "Compensation & Benefits — subrating"),
    ("senior_mgmt", "Senior Management — subrating"),
    ("wlb", "Work-Life Balance — subrating"),
    ("culture", "Culture & Values — subrating"),
    ("", ""),
    ("Election / RDD variables", ""),
    ("margin", "Union vote margin = vote_share - 0.5 (0 = 50% threshold)"),
    ("win", "1 if margin > 0 (union wins election); 0 otherwise"),
    ("post", "1 if review_date >= election_date; 0 otherwise"),
    ("abs_margin", "Absolute value of margin"),
    ("vote_share", "votes_for_union / (votes_for_union + votes_against_union)"),
    ("", ""),
    ("Interaction terms", ""),
    ("Win x Post", "Interaction of win and post (main review-level coefficient)"),
    ("Post x Margin", "Interaction of post and running variable"),
    ("Win x Post x Margin", "Triple interaction"),
    ("Win x Margin", "Interaction of win and running variable"),
    ("", ""),
    ("Event-level variables", ""),
    ("delta_y", "post-election mean rating - pre-election mean rating (event-level)"),
    ("pre_mean / post_mean", "Mean rating before/after election per election x outcome"),
    ("n_pre / n_post", "Number of reviews before/after election"),
    ("weight", "Harmonic mean: 2/(1/n_pre + 1/n_post), normalized to mean=1"),
    ("", ""),
    ("Sample filters", ""),
    ("current", "Current employees only (primary)"),
    ("all", "All employees (current + former, robustness)"),
    ("former", "Former employees only (diagnostic)"),
    ("Bandwidths", "Global (all margins); |m|<=0.20; |m|<=0.10"),
    ("Threshold", "pre>=N AND post>=N reviews per election"),
]
ws_vd.cell(row=1, column=1, value="Variable Definitions").font = title_font
r = 3
for var, defn in var_rows:
    if var == "" and defn == "":
        r += 1; continue
    if defn == "":
        ws_vd.cell(row=r, column=1, value=var).font = Font(bold=True, size=11)
    else:
        ws_vd.cell(row=r, column=1, value=var).font = Font(bold=True)
        ws_vd.cell(row=r, column=2, value=defn)
    style_row(ws_vd, r, 2)
    r += 1
auto_w(ws_vd)
ws_vd.column_dimensions['B'].width = 70

# Sheet: Table 2 — Review-Level DiD-RD Main Results
ws_t2 = wb.create_sheet("Table 2 Review DiD-RD")
ws_t2.cell(row=1, column=1, value=f"Table 2: Review-Level DiD-RD — {OUTCOMES[best_oc]} (Current Employees, +/-365d, pre>=3 post>=3)").font = title_font
ws_t2.merge_cells('A1:H1')

r = 3
ws_t2.cell(row=r, column=1, value="Dependent variable:").font = Font(italic=True)
ws_t2.cell(row=r, column=2, value=f"{OUTCOMES[best_oc]} (standardized)").font = Font(italic=True, bold=True)
r = 4
ws_t2.cell(row=r, column=1, value="Panel A: Without Election FE").font = Font(bold=True)
r = 5

# Build for one outcome across bandwidths
for panel_label, use_fe in [("Panel A: Without Election FE", False), ("Panel B: With Election FE (Win absorbed)", True)]:
    r += 1
    ws_t2.cell(row=r, column=1, value=panel_label).font = Font(bold=True)
    r += 1
    # Header
    headers = ["Variable"] + [f"({i+1}) {bw_label}" for i, (bw_label, _) in enumerate(BANDWIDTHS)]
    for c, h in enumerate(headers, 1):
        ws_t2.cell(row=r, column=c, value=h)
    style_hdr(ws_t2, r, len(headers))
    r += 1

    var_list = (["Win", "Post", "Win x Post", "Margin", "Win x Margin", "Post x Margin", "Win x Post x Margin"]
                if not use_fe else ["Win", "Post", "Win x Post", "Post x Margin", "Win x Post x Margin"])
    for vn in var_list:
        ws_t2.cell(row=r, column=1, value=vn)
        for j, (bw_label, _) in enumerate(BANDWIDTHS):
            spec = find_spec(df_rv_full, best_oc, "current", bw_label, election_fe=use_fe)
            col_num = 2 + j
            if vn == "Win" and use_fe:
                ws_t2.cell(row=r, column=col_num, value="absorbed").font = note_font
            elif spec is not None and f"{vn}_coef" in spec.columns:
                c_val = spec[f"{vn}_coef"].values[0]
                if pd.isna(c_val):
                    ws_t2.cell(row=r, column=col_num, value="—")
                else:
                    s_val = spec[f"{vn}_se"].values[0]
                    p = 2 * stats.t.sf(abs(c_val/s_val), df=1000) if s_val > 0 else np.nan
                    ws_t2.cell(row=r, column=col_num, value=f"{c_val:.3f}{stars(p)}")
                    # SE in next row
                    ws_t2.cell(row=r+1, column=col_num, value=f"({s_val:.3f})").font = note_font
            else:
                ws_t2.cell(row=r, column=col_num, value="—")
        style_row(ws_t2, r, len(headers))
        style_row(ws_t2, r+1, len(headers))
        r += 2

    # Bottom stats
    for label, col_name in [("Observations", "n_reviews"), ("Elections", "n_events"),
                              ("Firms (gvkey)", "n_gvkeys"), ("Win elections", "n_win_events"),
                              ("Loss elections", "n_loss_events")]:
        ws_t2.cell(row=r, column=1, value=label).font = Font(italic=True)
        for j, (bw_label, _) in enumerate(BANDWIDTHS):
            spec = find_spec(df_rv_full, best_oc, "current", bw_label, election_fe=use_fe)
            if spec is not None and col_name in spec.columns:
                ws_t2.cell(row=r, column=2+j, value=f"{int(spec[col_name].values[0]):,}")
        style_row(ws_t2, r, len(headers))
        r += 1
    r += 1

ws_t2.cell(row=r, column=1, value="Notes:").font = Font(italic=True, bold=True)
ws_t2.cell(row=r+1, column=1, value="*** p<0.01, ** p<0.05, * p<0.10. Standard errors in italics below coefficients.").font = note_font
ws_t2.cell(row=r+2, column=1, value="Coefficients in SD units. Employee filter: current. Window: +/-365 days. Threshold: pre>=3 post>=3.").font = note_font
ws_t2.cell(row=r+3, column=1, value="SE: HC1 robust. Year FE included in all columns.").font = note_font
auto_w(ws_t2)

# Sheet: Table 4 — Event-Level Delta RDD
ws_t4 = wb.create_sheet("Table 4 Event-Level RDD")
ws_t4.cell(row=1, column=1, value=f"Table 4: Event-Level Delta RDD — {OUTCOMES[best_oc]} (Current, +/-365d, pre>=1 post>=1, Weighted)").font = title_font
ws_t4.merge_cells('A1:G1')

r = 3
headers = ["Variable"] + [f"({i+1}) {bw_label}" for i, (bw_label, _) in enumerate(BANDWIDTHS)]
for c, h in enumerate(headers, 1):
    ws_t4.cell(row=r, column=c, value=h)
style_hdr(ws_t4, r, len(headers))
r += 1

for vn in ["Win", "Margin", "Win x Margin", "Constant"]:
    ws_t4.cell(row=r, column=1, value=vn)
    for j, (bw_label, _) in enumerate(BANDWIDTHS):
        spec = find_spec(df_ev_full, best_oc, "current", bw_label)
        if spec is not None and f"{vn}_coef" in spec.columns:
            c_val = spec[f"{vn}_coef"].values[0]
            if not pd.isna(c_val):
                s_val = spec[f"{vn}_se"].values[0]
                p = 2 * stats.t.sf(abs(c_val/s_val), df=1000) if s_val > 0 else np.nan
                ws_t4.cell(row=r, column=2+j, value=f"{c_val:.3f}{stars(p)}")
                ws_t4.cell(row=r+1, column=2+j, value=f"({s_val:.3f})").font = note_font
            else:
                ws_t4.cell(row=r, column=2+j, value="—")
        else:
            ws_t4.cell(row=r, column=2+j, value="—")
    style_row(ws_t4, r, len(headers))
    style_row(ws_t4, r+1, len(headers))
    r += 2

for label, col_name in [("Election events", "n_events"), ("Firms (gvkey)", "n_gvkeys"),
                          ("Win events", "n_win"), ("Loss events", "n_loss"),
                          ("Mean delta (losses)", "mean_delta_loss"), ("Mean delta (wins)", "mean_delta_win")]:
    ws_t4.cell(row=r, column=1, value=label).font = Font(italic=True)
    for j, (bw_label, _) in enumerate(BANDWIDTHS):
        spec = find_spec(df_ev_full, best_oc, "current", bw_label)
        if spec is not None and col_name in spec.columns:
            v = spec[col_name].values[0]
            ws_t4.cell(row=r, column=2+j, value=f"{v:.3f}" if isinstance(v, float) else f"{int(v):,}")
    style_row(ws_t4, r, len(headers))
    r += 1

ws_t4.cell(row=r+1, column=1, value="Notes: *** p<0.01, ** p<0.05, * p<0.10. SE in italics.").font = note_font
ws_t4.cell(row=r+2, column=1, value="Dependent variable: delta_y = post_mean - pre_mean, standardized. Weighted by harmonic mean of n_pre, n_post.").font = note_font
auto_w(ws_t4)

# Sheet: Table 5 — Bandwidth Robustness
ws_t5 = wb.create_sheet("Table 5 Bandwidth Robustness")
ws_t5.cell(row=1, column=1, value="Table 5: Bandwidth Robustness — Win x Post / Win across bandwidths (Current, +/-365d)").font = title_font
ws_t5.merge_cells('A1:G1')

r = 3
ws_t5.cell(row=r, column=1, value="Panel A: Review-Level DiD-RD — Win x Post").font = Font(bold=True)
r += 1
bw_all = [("Global", "global"), ("|m|<=0.30", "|m|<=0.30"), ("|m|<=0.20", "|m|<=0.20"),
          ("|m|<=0.10", "|m|<=0.10"), ("|m|<=0.05", "|m|<=0.05")]
ws_t5.cell(row=r, column=1, value="Outcome")
for j, (label, _) in enumerate(bw_all):
    ws_t5.cell(row=r, column=2+j, value=label)
style_hdr(ws_t5, r, len(bw_all)+1)
r += 1

for oc in OUTCOMES:
    ws_t5.cell(row=r, column=1, value=OUTCOMES[oc])
    for j, (_, bw_key) in enumerate(bw_all):
        spec = find_spec(df_rv_full, oc, "current", bw_key, election_fe=True)
        if spec is not None and "Win x Post_coef" in spec.columns:
            c = spec["Win x Post_coef"].values[0]
            s = spec["Win x Post_se"].values[0]
            if not pd.isna(c):
                p = 2 * stats.t.sf(abs(c/s), df=1000) if s > 0 else np.nan
                ws_t5.cell(row=r, column=2+j, value=f"{c:.3f}{stars(p)}")
                ws_t5.cell(row=r+1, column=2+j, value=f"({s:.3f})").font = note_font
            else:
                ws_t5.cell(row=r, column=2+j, value="—")
        else:
            ws_t5.cell(row=r, column=2+j, value="—")
    style_row(ws_t5, r, len(bw_all)+1)
    style_row(ws_t5, r+1, len(bw_all)+1)
    r += 2

r += 1
ws_t5.cell(row=r, column=1, value="Panel B: Event-Level Delta RDD — Win").font = Font(bold=True)
r += 1
ws_t5.cell(row=r, column=1, value="Outcome")
for j, (label, _) in enumerate(bw_all):
    ws_t5.cell(row=r, column=2+j, value=label)
style_hdr(ws_t5, r, len(bw_all)+1)
r += 1

for oc in OUTCOMES:
    ws_t5.cell(row=r, column=1, value=OUTCOMES[oc])
    for j, (_, bw_key) in enumerate(bw_all):
        spec = find_spec(df_ev_full, oc, "current", bw_key)
        if spec is not None and "Win_coef" in spec.columns:
            c = spec["Win_coef"].values[0]
            s = spec["Win_se"].values[0]
            if not pd.isna(c):
                p = 2 * stats.t.sf(abs(c/s), df=1000) if s > 0 else np.nan
                ws_t5.cell(row=r, column=2+j, value=f"{c:.3f}{stars(p)}")
                ws_t5.cell(row=r+1, column=2+j, value=f"({s:.3f})").font = note_font
            else:
                ws_t5.cell(row=r, column=2+j, value="—")
        else:
            ws_t5.cell(row=r, column=2+j, value="—")
    style_row(ws_t5, r, len(bw_all)+1)
    style_row(ws_t5, r+1, len(bw_all)+1)
    r += 2

ws_t5.cell(row=r+1, column=1, value="Notes: Coefficients with SE in parentheses. Panel A: election FE absorbed. Panel B: no FE (cross-sectional RDD).").font = note_font
auto_w(ws_t5)

# Sheet: Table A1 — All Employees
ws_a1 = wb.create_sheet("Table A1 All Employees")
ws_a1.cell(row=1, column=1, value="Table A1: Review-Level DiD-RD — All Employees").font = title_font
r = 3
ws_a1.cell(row=r, column=1, value="Outcome")
for j, (bw_label, _) in enumerate(BANDWIDTHS):
    ws_a1.cell(row=r, column=2+j*2, value=f"Win x Post ({bw_label})")
    ws_a1.cell(row=r, column=3+j*2, value="SE")
style_hdr(ws_a1, r, 7)
r += 1
for oc in OUTCOMES:
    ws_a1.cell(row=r, column=1, value=OUTCOMES[oc])
    for j, (bw_label, _) in enumerate(BANDWIDTHS):
        spec = find_spec(df_rv_full, oc, "all", bw_label, election_fe=True)
        if spec is not None and "Win x Post_coef" in spec.columns:
            c = spec["Win x Post_coef"].values[0]
            s = spec["Win x Post_se"].values[0]
            if not pd.isna(c):
                p_val = 2 * stats.t.sf(abs(c/s), df=1000) if s > 0 else np.nan
                ws_a1.cell(row=r, column=2+j*2, value=f"{c:.3f}{stars(p_val)}")
            else:
                ws_a1.cell(row=r, column=2+j*2, value="—")
            ws_a1.cell(row=r, column=3+j*2, value=f"({s:.3f})").font = note_font
    style_row(ws_a1, r, 7)
    r += 1
auto_w(ws_a1)

# Sheet: Table A2 — Former Employees
ws_a2 = wb.create_sheet("Table A2 Former Diagnostic")
ws_a2.cell(row=1, column=1, value="Table A2: Former Employees (Diagnostic Only) — Win x Post").font = title_font
r = 3
ws_a2.cell(row=r, column=1, value="Outcome")
for j, (bw_label, _) in enumerate(BANDWIDTHS):
    ws_a2.cell(row=r, column=2+j*2, value=f"({bw_label})")
style_hdr(ws_a2, r, 7)
r += 1
for oc in OUTCOMES:
    ws_a2.cell(row=r, column=1, value=OUTCOMES[oc])
    for j, (bw_label, _) in enumerate(BANDWIDTHS):
        spec = find_spec(df_rv_full, oc, "former", bw_label, election_fe=True)
        if spec is not None and "Win x Post_coef" in spec.columns:
            c = spec["Win x Post_coef"].values[0]
            s = spec["Win x Post_se"].values[0]
            if not pd.isna(c):
                p_val = 2 * stats.t.sf(abs(c/s), df=1000) if s > 0 else np.nan
                ws_a2.cell(row=r, column=2+j*2, value=f"{c:.3f}{stars(p_val)}")
                ws_a2.cell(row=r, column=3+j*2, value=f"({s:.3f})").font = note_font
    style_row(ws_a2, r, 7)
    r += 1
ws_a2.cell(row=r+1, column=1, value="DIAGNOSTIC ONLY. Do not use as main evidence.").font = Font(bold=True, color="CC0000")
auto_w(ws_a2)

# Sheet: Table A3 — All Outcomes (Current, |m|<=0.20)
ws_a3 = wb.create_sheet("Table 3 By Outcome")
ws_a3.cell(row=1, column=1, value="Table 3: Review-Level DiD-RD by Outcome (Current, +/-365d, |m|<=0.20, pre>=3 post>=3)").font = title_font
ws_a3.merge_cells('A1:G1')
r = 3
ws_a3.cell(row=r, column=1, value="Variable").font = hdr_font
for j, oc in enumerate(OUTCOMES):
    ws_a3.cell(row=r, column=2+j, value=OUTCOMES[oc]).font = hdr_font
style_hdr(ws_a3, r, len(OUTCOMES)+1)
r += 1

for vn in ["Win", "Post", "Win x Post"]:
    ws_a3.cell(row=r, column=1, value=vn)
    for j, oc in enumerate(OUTCOMES):
        spec = find_spec(df_rv_full, oc, "current", "|m|<=0.20", election_fe=True)
        if vn == "Win":
            ws_a3.cell(row=r, column=2+j, value="absorbed").font = note_font
        elif spec is not None and f"{vn}_coef" in spec.columns:
            c = spec[f"{vn}_coef"].values[0]
            s = spec[f"{vn}_se"].values[0]
            if not pd.isna(c):
                p_val = 2 * stats.t.sf(abs(c/s), df=1000) if s > 0 else np.nan
                ws_a3.cell(row=r, column=2+j, value=f"{c:.3f}{stars(p_val)}")
                ws_a3.cell(row=r+1, column=2+j, value=f"({s:.3f})").font = note_font
    style_row(ws_a3, r, len(OUTCOMES)+1)
    style_row(ws_a3, r+1, len(OUTCOMES)+1)
    r += 2

for label, col_name in [("Observations", "n_reviews"), ("Elections", "n_events")]:
    ws_a3.cell(row=r, column=1, value=label).font = Font(italic=True)
    for j, oc in enumerate(OUTCOMES):
        spec = find_spec(df_rv_full, oc, "current", "|m|<=0.20", election_fe=True)
        if spec is not None and col_name in spec.columns:
            ws_a3.cell(row=r, column=2+j, value=f"{int(spec[col_name].values[0]):,}")
    style_row(ws_a3, r, len(OUTCOMES)+1)
    r += 1

ws_a3.cell(row=r+1, column=1, value="Win absorbed by election FE. Current employees, |m|<=0.20, +/-365d, pre>=3 post>=3.").font = note_font
auto_w(ws_a3)

# Sheet: rdrobust (Table 6)
ws_rb = wb.create_sheet("Table 6 rdrobust")
ws_rb.cell(row=1, column=1, value="Table 6: Local-Linear RDD Robustness (Current, +/-365d, pre>=1 post>=1)").font = title_font
ws_rb.merge_cells('A1:G1')

# Load local-linear results
df_ll = pd.read_csv(PROJ / "outputs/rdd_rebuild/rdrobust_event_level_results.csv")
r = 3
ws_rb.cell(row=r, column=1, value="Outcome")
for j, h in enumerate(["tau", "SE", "p-value", "Bandwidth", "N effective", "N left", "N right"]):
    ws_rb.cell(row=r, column=2+j, value=h)
style_hdr(ws_rb, r, 8)
r += 1
for oc in OUTCOMES:
    s = df_ll[(df_ll["outcome"]==oc)&(df_ll["employee_filter"]=="current")&
              (df_ll["window_days"]==365)&(df_ll["threshold"]=="pre>=1_post>=1")]
    if len(s) > 0:
        b = s.iloc[(s["bandwidth"].astype(float)-0.20).abs().argsort().iloc[0]]
        sig = stars(b["p_value"])
        ws_rb.cell(row=r, column=1, value=OUTCOMES[oc])
        for j, k in enumerate(["tau","se","p_value","bandwidth","n_effective","n_left","n_right"]):
            v = b[k]
            ws_rb.cell(row=r, column=2+j, value=f"{v:.3f}{sig}" if k=="tau" else f"{v:.3f}" if isinstance(v,float) else int(v))
        style_row(ws_rb, r, 8)
        r += 1
ws_rb.cell(row=r+1, column=1, value="Triangular kernel, Silverman bandwidth. rdrobust Python package unavailable — manual implementation.").font = note_font
auto_w(ws_rb)

# Save workbook
wb_path = OUT / "union_glassdoor_rdd_paper_tables.xlsx"
wb.save(wb_path)
print(f"  Saved: union_glassdoor_rdd_paper_tables.xlsx")

# ═══════════════════════════════════════════════════════════════════════
# Save full coefficient data
# ═══════════════════════════════════════════════════════════════════════
df_rv_full.to_csv(OUT / "full_coefficient_outputs.csv", index=False)
print(f"  Saved: full_coefficient_outputs.csv ({len(df_rv_full)} review + event specs total)")

# ═══════════════════════════════════════════════════════════════════════
print(f"\n{'='*70}")
print("PAPER TABLES COMPLETE")
print(f"Outputs: {OUT}/")
print(f"  union_glassdoor_rdd_paper_tables.xlsx")
print(f"  union_glassdoor_rdd_paper_tables.md")
print(f"  full_coefficient_outputs.csv")
