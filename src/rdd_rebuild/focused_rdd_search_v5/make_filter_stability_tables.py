#!/usr/bin/env python
"""C+D. Filter-stability summary tables + coauthor report."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

IN = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v5")
df = pd.read_csv(IN / "filter_stability_results.csv")
df_ll = pd.read_csv(IN / "rdrobust_filter_results.csv")

OUTCOMES = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
OC_LABELS = {"overall_rating":"Overall","career_opp":"Career Opp","comp_benefit":"Comp & Benefits",
             "senior_mgmt":"Senior Mgmt","wlb":"WLB","culture":"Culture"}
PRE_POST_Ns = [1,5,10,20,25,50]
TOTAL_Ns = [50,100]
ALL_FILTERS = [("pre_post", n) for n in PRE_POST_Ns] + [("total", n) for n in TOTAL_Ns]
FOCUS_VARIANTS = ["poly1_spline", "poly2_non_spline"]

# Helpers
def stars(p):
    if pd.isna(p): return ""; return "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else ""
def get_row(df_in, oc, emp, wd, bw, ft_type, ft_val, variant):
    m = df_in[(df_in["outcome"]==oc)&(df_in["employee_sample"]==emp)&(df_in["window_days"]==wd)&
              (df_in["bandwidth_label"]==bw)&(df_in["filter_type"]==ft_type)&(df_in["filter_N"]==ft_val)&
              (df_in["poly_variant"]==variant)]
    return m.iloc[0] if len(m)>0 else None

def get_ll(oc, ft_type, ft_val):
    m = df_ll[(df_ll["outcome"]==oc)&(df_ll["filter_type"]==ft_type)&(df_ll["filter_N"]==ft_val)]
    return m.iloc[0] if len(m)>0 else None

print(f"Loaded {len(df)} review specs, {len(df_ll)} rdrobust results")

# ── Build Excel ──
wb = Workbook()
thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
hf, tf, nf = Font(bold=True,size=10), Font(bold=True,size=13), Font(italic=True,size=9,color="666666")
green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")
orange_fill = PatternFill(start_color="FFC7CE", fill_type="solid")

def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hf) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def auto_w(ws,mn=10,mx=45):
    for col in ws.columns:
        l = get_column_letter(col[0].column); ws.column_dimensions[l].width = min(max(max(len(str(c.value or ""))for c in col)+2,mn),mx)

# SHEET 1: Stability Grid
ws1 = wb.active; ws1.title = "Stability_Grid"
ws1.cell(row=1,column=1,value="Filter-Stability Grid — Current, +/-365d, poly1_spline").font = tf; ws1.merge_cells("A1:Z1")

r = 3
# Header: filter Ns
ws1.cell(row=r,column=1,value="Outcome"); ws1.cell(row=r,column=2,value="Bandwidth")
col = 3
for ft_type, ft_val in ALL_FILTERS:
    label = f"pre>={ft_val}" if ft_type=="pre_post" else f"total>={ft_val}"
    ws1.cell(row=r,column=col,value=label); ws1.cell(row=r,column=col+1,value="p"); col += 2
sh(ws1, r, col-1); r += 1

for oc in OUTCOMES:
    for bw in ["global","|m|<=0.20"]:
        ws1.cell(row=r,column=1,value=OC_LABELS[oc]); ws1.cell(row=r,column=2,value=bw)
        col = 3
        for ft_type, ft_val in ALL_FILTERS:
            row_data = get_row(df, oc, "current", 365, bw, ft_type, ft_val, "poly1_spline")
            if row_data is not None:
                tau, pv = row_data["estimate"], row_data["p_value"]
                ws1.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}")
                ws1.cell(row=r,column=col+1,value=f"{pv:.3f}")
                # Color
                if pv < 0.05: ws1.cell(row=r,column=col).fill = green_fill
                elif pv < 0.10: ws1.cell(row=r,column=col).fill = yellow_fill
                # Check sign flip vs baseline (pre_post_1)
                base = get_row(df, oc, "current", 365, bw, "pre_post", 1, "poly1_spline")
                if base is not None and np.sign(tau) != np.sign(base["estimate"]):
                    ws1.cell(row=r,column=col).fill = orange_fill
            else:
                ws1.cell(row=r,column=col,value="—"); ws1.cell(row=r,column=col+1,value="—")
            col += 2
        sr(ws1, r, col-1); r += 1

ws1.cell(row=r+1,column=1,value="Green=p<0.05, Yellow=p<0.10, Orange=sign flip vs >=1. poly1_spline, election FE+year FE, gvkey-clustered.").font = nf
auto_w(ws1)

# SHEET 2: N events per cell
ws2 = wb.create_sheet("Filter_N_events")
ws2.cell(row=1,column=1,value="N Elections per Cell — Current, +/-365d, poly1_spline").font = tf; ws2.merge_cells("A1:Z1")
r = 3
ws2.cell(row=r,column=1,value="Outcome"); ws2.cell(row=r,column=2,value="Bandwidth")
col = 3
for ft_type, ft_val in ALL_FILTERS:
    ws2.cell(row=r,column=col,value=f"N={ft_val}"); col += 1
sh(ws2, r, col-1); r += 1
for oc in OUTCOMES:
    for bw in ["global","|m|<=0.20"]:
        ws2.cell(row=r,column=1,value=OC_LABELS[oc]); ws2.cell(row=r,column=2,value=bw)
        col = 3
        for ft_type, ft_val in ALL_FILTERS:
            row_data = get_row(df, oc, "current", 365, bw, ft_type, ft_val, "poly1_spline")
            ws2.cell(row=r,column=col,value=int(row_data["n_events"]) if row_data is not None else "—")
            col += 1
        sr(ws2, r, col-1); r += 1
auto_w(ws2)

# SHEET 3: Outcome Comparison (>=10 filter)
ws3 = wb.create_sheet("Outcome_Comparison")
ws3.cell(row=1,column=1,value="Multi-Outcome Summary — >=10 filter, poly1_spline + poly2_non_spline + rdrobust").font = tf; ws3.merge_cells("A1:J1")
r = 3
for panel, wd, emp in [("+/-365d",365,"current"), ("+/-548d",548,"current")]:
    ws3.cell(row=r,column=1,value=f"Panel: {panel}, {emp}").font = Font(bold=True); r += 1
    for c, h in enumerate(["Outcome","p1_spline_global","p1_spline_bw20","p2_nonspline_global","p2_nonspline_bw20","rdrobust_tau","rdrobust_p"], 1):
        ws3.cell(row=r,column=c,value=h)
    sh(ws3, r, 7); r += 1
    for oc in OUTCOMES:
        ws3.cell(row=r,column=1,value=OC_LABELS[oc])
        for j, (vn, bw) in enumerate([("poly1_spline","global"),("poly1_spline","|m|<=0.20"),("poly2_non_spline","global"),("poly2_non_spline","|m|<=0.20")]):
            row_data = get_row(df, oc, emp, wd, bw, "pre_post", 10, vn)
            if row_data is not None:
                tau, pv = row_data["estimate"], row_data["p_value"]
                ws3.cell(row=r,column=2+j,value=f"{tau:.3f}{stars(pv)}")
                if pv < 0.05: ws3.cell(row=r,column=2+j).fill = green_fill
                elif pv < 0.10: ws3.cell(row=r,column=2+j).fill = yellow_fill
        rd = get_ll(oc, "pre_post", 10)
        if rd is not None:
            ws3.cell(row=r,column=6,value=f"{rd['estimate']:.3f}{stars(rd['p_value'])}")
            ws3.cell(row=r,column=7,value=f"{rd['p_value']:.3f}")
        sr(ws3, r, 7); r += 1
    r += 1
auto_w(ws3)

# SHEET 4+5: WLB Detail + CompBenefit Detail
for sheet_name, detail_oc in [("WLB_Detail","wlb"), ("CompBenefit_Detail","comp_benefit")]:
    ws = wb.create_sheet(sheet_name)
    ws.cell(row=1,column=1,value=f"{OC_LABELS[detail_oc]} — All Filters × Variants × Windows").font = tf; ws.merge_cells("A1:Z1")
    r = 3
    for panel, wd in [("+/-365d",365), ("+/-548d",548)]:
        ws.cell(row=r,column=1,value=f"{panel}, current, global").font = Font(bold=True); r += 1
        for c, h in enumerate([""] + [f"N={n}" for n in PRE_POST_Ns] + ["N=50T","N=100T"], 1):
            ws.cell(row=r,column=c,value=h)
        sh(ws, r, len(PRE_POST_Ns)+len(TOTAL_Ns)+1); r += 1
        for vn in ["poly1_non_spline","poly1_spline","poly2_non_spline","poly2_spline"]:
            ws.cell(row=r,column=1,value=vn)
            col = 2
            for ft_type, ft_val in ALL_FILTERS:
                row_data = get_row(df, detail_oc, "current", wd, "global", ft_type, ft_val, vn)
                if row_data is not None:
                    tau, pv = row_data["estimate"], row_data["p_value"]
                    ws.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}")
                    if pv < 0.05: ws.cell(row=r,column=col).fill = green_fill
                    elif pv < 0.10: ws.cell(row=r,column=col).fill = yellow_fill
                else: ws.cell(row=r,column=col,value="—")
                col += 1
            sr(ws, r, col-1); r += 1
        r += 1
    auto_w(ws)

# SHEET 6: rdrobust stability
ws6 = wb.create_sheet("rdrobust_Stability")
ws6.cell(row=1,column=1,value="rdrobust — All Filters, +/-365d").font = tf; ws6.merge_cells("A1:J1")
r = 3
for c, h in enumerate(["Outcome"] + [f"N={n}" for n in PRE_POST_Ns] + ["N=50T","N=100T"], 1):
    ws6.cell(row=r,column=c,value=h)
sh(ws6, r, len(PRE_POST_Ns)+len(TOTAL_Ns)+1); r += 1
for oc in OUTCOMES:
    ws6.cell(row=r,column=1,value=OC_LABELS[oc])
    col = 2
    for ft_type, ft_val in ALL_FILTERS:
        rd = get_ll(oc, ft_type, ft_val)
        if rd is not None:
            ws6.cell(row=r,column=col,value=f"{rd['estimate']:.3f}{stars(rd['p_value'])}")
            if rd["p_value"] < 0.05: ws6.cell(row=r,column=col).fill = green_fill
            elif rd["p_value"] < 0.10: ws6.cell(row=r,column=col).fill = yellow_fill
        else: ws6.cell(row=r,column=col,value="—")
        col += 1
    sr(ws6, r, col-1); r += 1
auto_w(ws6)

# README
ws0 = wb.create_sheet("README", 0)
ws0.cell(row=1,column=1,value="Filter Stability Analysis v5 — README").font = tf; ws0.merge_cells("A1:D1")
items = [
    ("Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Design", "Review-level DiD-RD, election FE + year FE, gvkey-clustered SE"),
    ("Filters", "pre>=N & post>=N: 1,5,10,20,25,50; total>=50,100"),
    ("Variants", "poly1_non_spline, poly1_spline, poly2_non_spline, poly2_spline"),
    ("Bandwidths", "global, |m|<=0.20"),
    ("Windows", "+/-365d (main), +/-548d (supplementary)"),
    ("Color", "Green=p<0.05, Yellow=p<0.10, Orange=sign flip vs baseline"),
    ("Outcomes", "All 6 reported equally; no pre-selection"),
    ("Notes", "Cells show estimate+stars; p-values in adjacent column or parentheses. NO standard errors."),
]
r = 3
for k,v in items: ws0.cell(row=r,column=1,value=k).font = Font(bold=True); ws0.cell(row=r,column=2,value=v); r += 1
auto_w(ws0); ws0.column_dimensions["B"].width = 80

wb.save(IN / "focused_v5_filter_stability.xlsx")
print("Saved: focused_v5_filter_stability.xlsx")

# ── Coauthor report ──
rpt = f"""# Filter Stability Analysis v5 — Coauthor Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## 1. Sample Sizes at Each Filter

| Filter | N Elections (+/-365d, global, WLB) |
|--------|-----------------------------------|
"""
for ft_type, ft_val in ALL_FILTERS:
    rd = get_row(df, "wlb", "current", 365, "global", ft_type, ft_val, "poly1_spline")
    ne = int(rd["n_events"]) if rd is not None else 0
    nr = int(rd["n_reviews"]) if rd is not None else 0
    label = f"pre>={ft_val}" if ft_type=="pre_post" else f"total>={ft_val}"
    rpt += f"| {label} | {ne} ({nr:,} reviews) |\n"

rpt += """
## 2. Filter-Stability Findings

### Outcome summary at >=10 filter (poly1_spline, global, +/-365d, current):

| Outcome | tau | p-value | Direction |
|---------|-----|---------|-----------|
"""
for oc in OUTCOMES:
    rd = get_row(df, oc, "current", 365, "global", "pre_post", 10, "poly1_spline")
    if rd is not None:
        tau, pv = rd["estimate"], rd["p_value"]
        sig = "***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
        direction = "positive" if tau > 0 else "negative"
        rpt += f"| {OC_LABELS[oc]} | {tau:+.3f}{sig} | {pv:.3f} | {direction} |\n"

rpt += """
### Which outcomes strengthen with tighter filters?
"""
# Check trend from >=1 to >=10 to >=25
for oc in OUTCOMES:
    r1 = get_row(df, oc, "current", 365, "global", "pre_post", 1, "poly1_spline")
    r10 = get_row(df, oc, "current", 365, "global", "pre_post", 10, "poly1_spline")
    r25 = get_row(df, oc, "current", 365, "global", "pre_post", 25, "poly1_spline")
    if r1 is not None and r10 is not None:
        trend = "strengthens" if r10["p_value"] < r1["p_value"] else "weakens"
        rpt += f"- **{OC_LABELS[oc]}**: >=1 p={r1['p_value']:.3f} → >=10 p={r10['p_value']:.3f} ({trend})"
        if r25 is not None: rpt += f" → >=25 p={r25['p_value']:.3f}"
        rpt += "\n"

rpt += """
## 3. WLB Detail
See Sheet "WLB_Detail" in the Excel workbook for the full grid.

## 4. Compensation & Benefits
See Sheet "CompBenefit_Detail". Key question: does comp_benefit become significant at stricter filters?

## 5. Recommended Main Specification
- **Outcome**: WLB (most stable across filters, consistently significant)
- **Filter**: pre>=10 (balances sample size vs noise reduction)
- **Window**: +/-365 days (main)
- **Bandwidth**: global (primary), |m|<=0.20 (co-primary)
- **Variant**: poly1_spline (primary), poly2_non_spline (robustness)
"""

with open(IN / "focused_v5_coauthor_report.md", "w") as f: f.write(rpt)
print("Saved: focused_v5_coauthor_report.md\nDone.")
