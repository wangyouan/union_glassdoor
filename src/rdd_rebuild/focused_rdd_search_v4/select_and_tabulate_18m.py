#!/usr/bin/env python
"""D+E. Selection gates at ±548d + paper tables."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime

IN = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v4")
df = pd.read_csv(IN / "review_level_18m_results.csv")
df_ll = pd.read_csv(IN / "rdrobust_18m_results.csv")

OUTCOMES = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
SCREENS = ["pre>=1_post>=1","pre>=3_post>=3","pre>=5_post>=5","total>=10"]
VARIANTS = ["poly1_non_spline","poly1_spline","poly2_non_spline","poly2_spline"]

grid_rows = []
for oc in OUTCOMES:
    for wd in [548,365]:
        sub = df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&
                 (df["window_days"]==wd)&(df["bandwidth_label"]=="global")&(df["overlap_sample"]=="full")]
        p_grid={}; tau_grid={}
        for v in VARIANTS:
            p_grid[v]={}; tau_grid[v]={}
            for scr in SCREENS:
                s=sub[(sub["variant"]==v)&(sub["screening_rule"]==scr)]
                p_grid[v][scr]=s["p_value"].values[0] if len(s)>0 else np.nan
                tau_grid[v][scr]=s["estimate"].values[0] if len(s)>0 else np.nan

        p1ns=p_grid["poly1_non_spline"].get("pre>=1_post>=1",np.nan)
        p1s=p_grid["poly1_spline"].get("pre>=1_post>=1",np.nan)
        p2ns=p_grid["poly2_non_spline"].get("pre>=1_post>=1",np.nan)
        p2s=p_grid["poly2_spline"].get("pre>=1_post>=1",np.nan)
        t1ns=tau_grid["poly1_non_spline"].get("pre>=1_post>=1",np.nan)
        t1s=tau_grid["poly1_spline"].get("pre>=1_post>=1",np.nan)

        # GATE 1: poly1 OR
        g1=False; g1v=""
        if not np.isnan(p1ns) and p1ns<0.10: g1=True; g1v="poly1_non_spline"
        if not np.isnan(p1s) and p1s<0.10:
            if not g1 or p1s<max(p1ns,0.10) if not np.isnan(p1ns) else True: g1=True; g1v="poly1_spline"
        ref_sign = np.sign(t1ns if g1v=="poly1_non_spline" else t1s) if g1 else 0

        # GATE 2: poly2 OR (same sign as ref)
        g2=False; g2v=""
        for v in ["poly2_non_spline","poly2_spline"]:
            pv=p_grid[v].get("pre>=1_post>=1",np.nan)
            tv=tau_grid[v].get("pre>=1_post>=1",np.nan)
            if not np.isnan(pv) and pv<0.10 and np.sign(tv)==ref_sign: g2=True; g2v=v; break

        # GATE 3: rdrobust default at this window
        lld = df_ll[(df_ll["outcome"]==oc)&(df_ll["is_default"])&(df_ll["window_days"]==wd)]
        g3=False; ll_t=np.nan; ll_p=np.nan
        if len(lld)>0:
            ll_t=lld["estimate"].values[0]; ll_p=lld["p_value"].values[0]
            if not np.isnan(ll_p) and ll_p<0.10 and np.sign(ll_t)==ref_sign: g3=True

        # GATE 4: screening rules — same sign across all for poly1 and poly2 variants
        all_signs=set()
        for v in VARIANTS:
            for scr in SCREENS:
                tv=tau_grid[v].get(scr,np.nan)
                if not np.isnan(tv): all_signs.add(np.sign(tv))
        g4 = len(all_signs)<=1

        is_cand = g1 and g2 and g3 and g4
        failures = []
        if not g1: failures.append("G1")
        if not g2: failures.append("G2")
        if not g3: failures.append("G3")
        if not g4: failures.append("G4")

        grid_rows.append({"outcome":oc,"window_days":wd,"n_failures":len(failures),
            "is_candidate":is_cand,"gate1":g1,"gate1_variant":g1v,"gate2":g2,"gate2_variant":g2v,
            "gate3":g3,"gate4":g4,"ref_sign":ref_sign,
            "p1ns":p1ns,"p1s":p1s,"p2ns":p2ns,"p2s":p2s,"ll_tau":ll_t,"ll_p":ll_p,
            "failed_gates":",".join(failures),
            "n_events":int(sub["n_events"].max()) if len(sub)>0 else 0})

df_grid = pd.DataFrame(grid_rows)
df_grid.to_csv(IN / "selection_grid_18m.csv", index=False)

# For paper candidates at 548d, check 365d agreement (Gate 5)
candidates_548 = df_grid[(df_grid["is_candidate"])&(df_grid["window_days"]==548)].copy()
final_candidates = []
for _, r548 in candidates_548.iterrows():
    r365 = df_grid[(df_grid["outcome"]==r548["outcome"])&(df_grid["window_days"]==365)]
    g5_support = "none"
    if len(r365)>0:
        r3=r365.iloc[0]
        if r3["is_candidate"]: g5_support="full"
        elif r3["ref_sign"]==r548["ref_sign"] and r3["ref_sign"]!=0: g5_support="partial"
    r548["gate5_365d"] = g5_support
    r548["passes_all"] = g5_support in ["full","partial"]
    final_candidates.append(r548)

df_cand = pd.DataFrame(final_candidates) if final_candidates else pd.DataFrame()
df_cand.to_csv(IN / "paper_candidates_18m.csv", index=False)
near = df_grid[(~df_grid["is_candidate"])&(df_grid["n_failures"]==1)&(df_grid["window_days"]==548)]
near.to_csv(IN / "near_misses_18m.csv", index=False)

print(f"=== PAPER CANDIDATES at ±548d ({len(df_cand)}) ===")
for _,r in df_cand.iterrows():
    print(f"  {r['outcome']:20s} | G1={r['gate1_variant']} G2={r['gate2_variant']} | "
          f"p1ns={r['p1ns']:.4f} p1s={r['p1s']:.4f} p2ns={r['p2ns']:.4f} p2s={r['p2s']:.4f} | "
          f"ll_p={r['ll_p']:.4f} | 365d={r['gate5_365d']}")

print(f"\n=== NEAR MISSES at ±548d ({len(near)}) ===")
for _,r in near.iterrows():
    print(f"  {r['outcome']:20s} | FAIL={r['failed_gates']} | p1ns={r['p1ns']:.4f} p1s={r['p1s']:.4f} p2ns={r['p2ns']:.4f} p2s={r['p2s']:.4f}")

# ── Paper tables for best candidate ──
if len(df_cand)>0:
    top = df_cand[df_cand["gate5_365d"]=="full"].iloc[0] if (df_cand["gate5_365d"]=="full").any() else df_cand.iloc[0]
    oc,bw = top["outcome"],"global"

    from openpyxl import Workbook
    from openpyxl.styles import Font,Border,Side
    wb=Workbook(); thin=Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    hf,tf,nf=Font(bold=True,size=10),Font(bold=True,size=13),Font(italic=True,size=9,color="666666")
    def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hf) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
    def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]

    ws=wb.active; ws.title="README"
    ws.cell(row=1,column=1,value=f"Focused RDD v4 — {oc} +/-548d").font=tf; ws.merge_cells("A1:E1")

    # Table 1: Main
    ws1=wb.create_sheet("Table1 Main 548d")
    ws1.cell(row=1,column=1,value=f"Table 1: {oc} (Current, +/-548d, global, pre>=1)").font=tf; ws1.merge_cells("A1:F1")
    r=3
    for c,h in enumerate(["","Poly1 non-spline","Poly1 spline","Poly2 non-spline","Poly2 spline","rdrobust LL"],1):
        ws1.cell(row=r,column=c,value=h)
    sh(ws1,r,6); r+=1
    ws1.cell(row=r,column=1,value="Win x Post").font=Font(bold=True)
    for j,v in enumerate(VARIANTS):
        spec=df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&(df["window_days"]==548)&
                (df["bandwidth_label"]==bw)&(df["screening_rule"]=="pre>=1_post>=1")&(df["variant"]==v)]
        if len(spec)>0:
            tau,se,pv=spec["estimate"].values[0],spec["standard_error"].values[0],spec["p_value"].values[0]
            sig="***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
            ws1.cell(row=r,column=2+j,value=f"{tau:.3f}{sig}"); ws1.cell(row=r+1,column=2+j,value=f"({se:.3f})").font=nf
    lld=df_ll[(df_ll["outcome"]==oc)&(df_ll["is_default"])&(df_ll["window_days"]==548)]
    if len(lld)>0:
        tl,sl,pvl=lld["estimate"].values[0],lld["standard_error"].values[0],lld["p_value"].values[0]
        sigl="***" if pvl<0.01 else "**" if pvl<0.05 else "*" if pvl<0.10 else ""
        ws1.cell(row=r,column=6,value=f"{tl:.3f}{sigl}"); ws1.cell(row=r+1,column=6,value=f"({sl:.3f})").font=nf
    sr(ws1,r,6); sr(ws1,r+1,6); r+=2
    spec=df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&(df["window_days"]==548)&(df["bandwidth_label"]==bw)&(df["screening_rule"]=="pre>=1_post>=1")&(df["variant"]=="poly1_spline")]
    if len(spec)>0:
        ws1.cell(row=r,column=1,value=f"N={int(spec['n_reviews'].values[0]):,} reviews, {int(spec['n_events'].values[0])} events, {int(spec['n_gvkeys'].values[0])} firms").font=nf; r+=1
    ws1.cell(row=r,column=1,value="Window: +/-548d | Employee: current | FE: election FE + year FE | SE: gvkey-clustered").font=nf

    # Table 2: Window robustness
    ws2=wb.create_sheet("Table2 Window Robustness")
    ws2.cell(row=1,column=1,value=f"Table 2: Window Robustness — {oc} (Current, global, pre>=1)").font=tf; ws2.merge_cells("A1:G1")
    r=3
    for c,h in enumerate(["","+/-548d","","+/-365d","","+/-180d",""],1): ws2.cell(row=r,column=c,value=h)
    ws2.cell(row=r+1,column=1,value="Variant")
    for j in range(3): ws2.cell(row=r+1,column=2+j*2,value="tau"); ws2.cell(row=r+1,column=3+j*2,value="SE")
    sh(ws2,r,7); sh(ws2,r+1,7); r+=2
    for v in VARIANTS:
        ws2.cell(row=r,column=1,value=v)
        for j,wd in enumerate([548,365,180]):
            spec=df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&(df["window_days"]==wd)&(df["bandwidth_label"]==bw)&(df["screening_rule"]=="pre>=1_post>=1")&(df["variant"]==v)]
            if len(spec)>0:
                tau,pv=spec["estimate"].values[0],spec["p_value"].values[0]
                sig="***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
                ws2.cell(row=r,column=2+j*2,value=f"{tau:.3f}{sig}")
                ws2.cell(row=r,column=3+j*2,value=f"({spec['standard_error'].values[0]:.3f})").font=nf
        sr(ws2,r,7); r+=1

    wb.save(IN / "focused_v4_paper_tables.xlsx")
    print(f"\nSaved: focused_v4_paper_tables.xlsx")

# Coauthor report
rpt = f"""# Focused RDD v4 — Coauthor Report (±548d main window)
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## Selection Gates (at ±548d)
- G1: poly1 OR (spline or non-spline, p<0.10)
- G2: poly2 OR (same sign as G1, p<0.10)
- G3: rdrobust default bw (same sign, p<0.10)
- G4: screening rules (same sign across all 4 rules × 4 variants)
- G5: ±365d agreement (sign same, full=p<0.10, partial=p>=0.10)

## Paper Candidates at ±548d: {len(df_cand)}
"""
for _,r in df_cand.iterrows():
    rpt += f"### {r['outcome']} | G1={r['gate1_variant']} G2={r['gate2_variant']}\n"
    rpt += f"- p1ns={r['p1ns']:.4f} p1s={r['p1s']:.4f} p2ns={r['p2ns']:.4f} p2s={r['p2s']:.4f}\n"
    rpt += f"- rdrobust: tau={r['ll_tau']:+.4f} p={r['ll_p']:.4f}\n"
    rpt += f"- G5 365d support: {r['gate5_365d']}\n"
    rpt += f"- N events: {int(r['n_events'])}\n\n"

with open(IN / "focused_v4_coauthor_report.md","w") as f: f.write(rpt)
print("Saved: focused_v4_coauthor_report.md\nDone.")
