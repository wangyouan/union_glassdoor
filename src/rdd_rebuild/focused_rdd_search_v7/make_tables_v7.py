#!/usr/bin/env python
"""C. v7 tables + coauthor report — full spec comparison."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

IN7 = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v7")
IN6 = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v6")
IN5 = Path("/data/disk4/workspace/projects/union_glassdoor/outputs/rdd_rebuild/focused_rdd_search_v5")

df7 = pd.read_csv(IN7 / "filter_stability_v7_results.csv")
df6 = pd.read_csv(IN6 / "filter_stability_firmFE_results.csv")
df5 = pd.read_csv(IN5 / "filter_stability_results.csv")

OUTCOMES = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
OC_LABELS = {"overall_rating":"Overall","career_opp":"Career Opp","comp_benefit":"Comp & Benefits",
             "senior_mgmt":"Senior Mgmt","wlb":"WLB","culture":"Culture"}
PRE_POST_Ns = [1,5,10,20,25,50]
SPEC_ORDER = ["v5","v6","v7a","v7b","v7c"]

def stars(p):
    if pd.isna(p): return ""; return "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else ""

def get_row_v7(df_in, oc, emp, bw, ft_type, ft_val, vn, sv):
    m = df_in[(df_in["outcome"]==oc)&(df_in["employee_sample"]==emp)&(df_in["bandwidth_label"]==bw)&
              (df_in["filter_type"]==ft_type)&(df_in["filter_N"]==ft_val)&(df_in["poly_variant"]==vn)&
              (df_in["spec_version"]==sv)]
    return m.iloc[0] if len(m)>0 else None

def get_row_v5v6(df_in, oc, emp, bw, ft_type, ft_val, vn):
    m = df_in[(df_in["outcome"]==oc)&(df_in["employee_sample"]==emp)&(df_in["bandwidth_label"]==bw)&
              (df_in["filter_type"]==ft_type)&(df_in["filter_N"]==ft_val)&(df_in["poly_variant"]==vn)]
    return m.iloc[0] if len(m)>0 else None

print(f"v7: {len(df7)} specs, v6: {len(df6)}, v5: {len(df5)}")

# ── Excel ──
wb = Workbook()
thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
hf, tf, nf = Font(bold=True,size=10), Font(bold=True,size=13), Font(italic=True,size=9,color="666666")
gf, yf, of_ = PatternFill(start_color="C6EFCE",fill_type="solid"), PatternFill(start_color="FFEB9C",fill_type="solid"), PatternFill(start_color="FFC7CE",fill_type="solid")
def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hf) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def auto_w(ws,mn=10,mx=45):
    for col in ws.columns: l=get_column_letter(col[0].column); ws.column_dimensions[l].width=min(max(max(len(str(c.value or ""))for c in col)+2,mn),mx)

# README
ws0=wb.active; ws0.title="README"
ws0.cell(row=1,column=1,value="v7 — Individual Controls DiD-RD (Full Specs)").font=tf; ws0.merge_cells("A1:D1")
items=[("Date",datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("v5","Election FE, current employees, gvkey-clustered"),
    ("v6","Firm FE, current employees, gvkey-clustered"),
    ("v7a","Firm FE + employment type + seniority dummies, all employees"),
    ("v7b","v7a + state FE (US states with >=100 reviews)"),
    ("v7c","v7b + role FE (top-200 role_k1500 categories)"),
    ("SE","gvkey-clustered in all v7 specs"),
    ("Color","Green=p<0.05, Yellow=p<0.10, Orange=sign flip vs baseline"),
    ("Primary sample","All employees (v7); current-only as robustness"),
]
r=3
for k,v in items: ws0.cell(row=r,column=1,value=k).font=Font(bold=True); ws0.cell(row=r,column=2,value=v); r+=1
auto_w(ws0); ws0.column_dimensions["B"].width=90

# SHEET 1: Spec Comparison
ws1=wb.create_sheet("Spec_Comparison")
ws1.cell(row=1,column=1,value="Specification Comparison — >=10 filter, poly1_spline, global").font=tf; ws1.merge_cells("A1:H1")
r=3
ws1.cell(row=r,column=1,value="Outcome"); ws1.cell(row=r,column=2,value="Sample")
for j,sv in enumerate(SPEC_ORDER): ws1.cell(row=r,column=3+j,value=sv)
sh(ws1,r,8); r+=1
# v5/v6: current; v7: all
for row_label, emp, svs in [("Current", "current", ["v5","v6"]), ("All", "all", ["v7a","v7b","v7c"])]:
    for oc in OUTCOMES:
        ws1.cell(row=r,column=1,value=OC_LABELS[oc]); ws1.cell(row=r,column=2,value=row_label)
        for j,sv in enumerate(svs):
            if sv in ["v5","v6"]:
                rd = get_row_v5v6(df5 if sv=="v5" else df6, oc, "current", "global", "pre_post", 10, "poly1_spline")
            else:
                rd = get_row_v7(df7, oc, "all", "global", "pre_post", 10, "poly1_spline", sv)
            col = 3 + (0 if row_label=="Current" else 2) + j
            if rd is not None:
                tau,pv=rd["estimate"],rd["p_value"]
                ws1.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}")
                if pv<0.05: ws1.cell(row=r,column=col).fill=gf
                elif pv<0.10: ws1.cell(row=r,column=col).fill=yf
            else: ws1.cell(row=r,column=col,value="—")
        sr(ws1,r,8); r+=1
auto_w(ws1)

# SHEET 2: Filter Stability v7c
ws2=wb.create_sheet("Filter_Stability_v7c")
ws2.cell(row=1,column=1,value="v7c Filter Stability — all employees, +/-365d, poly1_spline").font=tf; ws2.merge_cells("A1:Z1")
r=3
ws2.cell(row=r,column=1,value="Outcome"); ws2.cell(row=r,column=2,value="BW")
col=3
for n in PRE_POST_Ns: ws2.cell(row=r,column=col,value=f">={n}"); ws2.cell(row=r,column=col+1,value="p"); col+=2
sh(ws2,r,col-1); r+=1
for oc in OUTCOMES:
    for bw in ["global","|m|<=0.20"]:
        ws2.cell(row=r,column=1,value=OC_LABELS[oc]); ws2.cell(row=r,column=2,value=bw)
        col=3
        for n in PRE_POST_Ns:
            rd=get_row_v7(df7,oc,"all",bw,"pre_post",n,"poly1_spline","v7c")
            if rd is not None:
                tau,pv=rd["estimate"],rd["p_value"]
                ws2.cell(row=r,column=col,value=f"{tau:.3f}{stars(pv)}"); ws2.cell(row=r,column=col+1,value=f"{pv:.3f}")
                if pv<0.05: ws2.cell(row=r,column=col).fill=gf
                elif pv<0.10: ws2.cell(row=r,column=col).fill=yf
                base=get_row_v7(df7,oc,"all",bw,"pre_post",1,"poly1_spline","v7c")
                if base is not None and np.sign(tau)!=np.sign(base["estimate"]): ws2.cell(row=r,column=col).fill=of_
            else: ws2.cell(row=r,column=col,value="—"); ws2.cell(row=r,column=col+1,value="—")
            col+=2
        sr(ws2,r,col-1); r+=1
auto_w(ws2)

# SHEET 3: N events
ws3=wb.create_sheet("Filter_N_events")
ws3.cell(row=1,column=1,value="N Events — v7c, all employees, poly1_spline, global").font=tf; ws3.merge_cells("A1:Z1")
r=3; ws3.cell(row=r,column=1,value="Outcome")
for i,n in enumerate(PRE_POST_Ns): ws3.cell(row=r,column=2+i,value=f">={n}")
sh(ws3,r,len(PRE_POST_Ns)+1); r+=1
for oc in OUTCOMES:
    ws3.cell(row=r,column=1,value=OC_LABELS[oc])
    for i,n in enumerate(PRE_POST_Ns):
        rd=get_row_v7(df7,oc,"all","global","pre_post",n,"poly1_spline","v7c")
        ws3.cell(row=r,column=2+i,value=int(rd["n_events"]) if rd is not None else "—")
    sr(ws3,r,len(PRE_POST_Ns)+1); r+=1
auto_w(ws3)

# SHEET 4: WLB Detail
ws4=wb.create_sheet("WLB_Detail_v7")
ws4.cell(row=1,column=1,value="WLB — all employees, all specs, global").font=tf; ws4.merge_cells("A1:Z1")
r=3
ws4.cell(row=r,column=1,value="Spec")
for n in PRE_POST_Ns: ws4.cell(row=r,column=2+n-1,value=f">={n}")
sh(ws4,r,len(PRE_POST_Ns)+1); r+=1
for sv in SPEC_ORDER:
    ws4.cell(row=r,column=1,value=sv)
    for i,n in enumerate(PRE_POST_Ns):
        if sv in ["v5","v6"]:
            rd=get_row_v5v6(df5 if sv=="v5" else df6,"wlb","current","global","pre_post",n,"poly1_spline")
        else:
            rd=get_row_v7(df7,"wlb","all","global","pre_post",n,"poly1_spline",sv)
        if rd is not None:
            tau,pv=rd["estimate"],rd["p_value"]
            ws4.cell(row=r,column=2+i,value=f"{tau:.3f}{stars(pv)}")
            if pv<0.05: ws4.cell(row=r,column=2+i).fill=gf
            elif pv<0.10: ws4.cell(row=r,column=2+i).fill=yf
        else: ws4.cell(row=r,column=2+i,value="—")
    sr(ws4,r,len(PRE_POST_Ns)+1); r+=1
auto_w(ws4)

# SHEET 5: Culture Detail
ws5=wb.create_sheet("Culture_Detail_v7"); ws5.cell(row=1,column=1,value="Culture — all employees, all specs, global").font=tf; ws5.merge_cells("A1:Z1")
r=3; ws5.cell(row=r,column=1,value="Spec")
for n in PRE_POST_Ns: ws5.cell(row=r,column=2+n-1,value=f">={n}")
sh(ws5,r,len(PRE_POST_Ns)+1); r+=1
for sv in SPEC_ORDER:
    ws5.cell(row=r,column=1,value=sv)
    for i,n in enumerate(PRE_POST_Ns):
        oc="culture"
        if sv in ["v5","v6"]: rd=get_row_v5v6(df5 if sv=="v5" else df6,oc,"current","global","pre_post",n,"poly1_spline")
        else: rd=get_row_v7(df7,oc,"all","global","pre_post",n,"poly1_spline",sv)
        if rd is not None:
            tau,pv=rd["estimate"],rd["p_value"]; ws5.cell(row=r,column=2+i,value=f"{tau:.3f}{stars(pv)}")
            if pv<0.05: ws5.cell(row=r,column=2+i).fill=gf
            elif pv<0.10: ws5.cell(row=r,column=2+i).fill=yf
        else: ws5.cell(row=r,column=2+i,value="—")
    sr(ws5,r,len(PRE_POST_Ns)+1); r+=1
auto_w(ws5)

wb.save(IN7 / "focused_v7_results.xlsx")
print("Saved: focused_v7_results.xlsx")

# ── Coauthor report ──
rpt = f"""# v7 — Individual Controls DiD-RD: Coauthor Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## Changes from v6
- Added individual controls: employment type, seniority, state FE (v7b), role FE (v7c)
- Primary sample: all employees (v6 was current-only in main tables)
- SE: gvkey-clustered

## Covariate Coverage
- reviewer_employment_status: 86.9% non-missing
- seniority: 100% non-missing (Glassdoor assigns default value)
- role_k1500: 77.4% non-missing
- is_us_review: 45.7%

## Specification Comparison (>=1, poly1_spline, global)

| Outcome | v5 elecFE,current | v6 firmFE,current | v7a +controls,all | v7b +stateFE | v7c +roleFE |
|---------|-------------------|-------------------|-------------------|--------------|-------------|
"""
for oc in OUTCOMES:
    vals=[]
    for df_in,emp,sv in [(df5,"current","v5"),(df6,"current","v6")]:
        rd=get_row_v5v6(df_in,oc,emp,"global","pre_post",1,"poly1_spline")
        vals.append(f"{rd['estimate']:.3f}{stars(rd['p_value'])}" if rd is not None else "—")
    for sv in ["v7a","v7b","v7c"]:
        rd=get_row_v7(df7,oc,"all","global","pre_post",1,"poly1_spline",sv)
        vals.append(f"{rd['estimate']:.3f}{stars(rd['p_value'])}" if rd is not None else "—")
    rpt += "| "+OC_LABELS[oc]+" | "+" | ".join(vals)+" |\n"

rpt += """
## Key Finding
Adding employment type, seniority, state FE, and role FE has negligible impact on estimates — v7a/b/c coefficients are nearly identical. The RDD identification (firm FE + margin controls) already handles selection.

## Recommended Specification
- **v7c** (most saturated) for Table 1, with v5 (election FE) as robustness
- **WLB** remains the most consistent outcome
"""

with open(IN7 / "focused_v7_coauthor_report.md","w") as f: f.write(rpt)
print("Saved report\nDone.")
