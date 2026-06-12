#!/usr/bin/env python
"""D. Combine all focused RDD results and produce consistency summary."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime

PROJ = Path("/data/disk4/workspace/projects/union_glassdoor")
IN_DIR = PROJ / "outputs/rdd_rebuild/focused_rdd_search"

print("Loading results...")
df_rv = pd.read_csv(IN_DIR / "review_level_focused_rdd_results.csv")
df_fy = pd.read_csv(IN_DIR / "firm_year_aggregate_rdd_results.csv")
df_fq = pd.read_csv(IN_DIR / "firm_quarter_aggregate_rdd_results.csv")

df_rv["framework"] = "review_level"
df_fy["framework"] = "firm_year"
df_fq["framework"] = "firm_quarter"

# Combine
all_cols = ["framework","outcome","outcome_label","employee_sample","window_days","bandwidth_label",
            "threshold_rule","fixed_effects","polynomial_order","min_comment_rule",
            "n_observations" if "n_observations" in df_rv.columns else "n_reviews",
            "n_events","n_gvkeys","n_win_events","n_loss_events",
            "coefficient_of_interest","estimate","standard_error","t_stat","p_value"]
# Normalize names
df_rv["n_observations"] = df_rv["n_reviews"]; df_rv["min_comment_rule"] = "pre>=1_post>=1"
df_rv["polynomial_order"] = 1; df_fy["threshold_rule"] = df_fy["min_comment_rule"]

all_df = pd.concat([df_rv, df_fy, df_fq], ignore_index=True)
all_df.to_csv(IN_DIR / "focused_rdd_all_results.csv", index=False)
print(f"Combined: {len(all_df)} total specifications")

# Consistency assessment
OUTCOMES_ORDER = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
summary_rows = []
for oc in OUTCOMES_ORDER:
    for emp in ["current","all"]:
        sub = all_df[(all_df["outcome"]==oc)&(all_df["employee_sample"]==emp)]
        if len(sub) < 5: continue

        est = sub["estimate"].dropna()
        pos_share = (est > 0).mean(); neg_share = (est < 0).mean()
        sig10 = (sub["p_value"] < 0.10).sum()
        sig5 = (sub["p_value"] < 0.05).sum()
        sig1 = (sub["p_value"] < 0.01).sum()
        median_est = est.median(); mean_est = est.mean()
        iqr = est.quantile(0.75) - est.quantile(0.25)

        # By framework
        rv_sub = sub[sub["framework"]=="review_level"]
        fy_sub = sub[sub["framework"]=="firm_year"]
        fq_sub = sub[sub["framework"]=="firm_quarter"]
        rv_sign = np.sign(rv_sub["estimate"].median()) if len(rv_sub)>0 else 0
        fy_sign = np.sign(fy_sub["estimate"].median()) if len(fy_sub)>0 else 0
        fq_sign = np.sign(fq_sub["estimate"].median()) if len(fq_sub)>0 else 0

        # Cross-framework consistency
        signs = [s for s in [rv_sign, fy_sign, fq_sign] if s != 0]
        frameworks_agree = len(set(signs)) <= 1 if len(signs) >= 2 else True

        # Current vs all
        cur_sub = all_df[(all_df["outcome"]==oc)&(all_df["employee_sample"]=="current")]
        all_sub = all_df[(all_df["outcome"]==oc)&(all_df["employee_sample"]=="all")]
        cur_sign = np.sign(cur_sub["estimate"].median()) if len(cur_sub)>0 else 0
        all_sign = np.sign(all_sub["estimate"].median()) if len(all_sub)>0 else 0
        cur_all_agree = (cur_sign == all_sign) if cur_sign != 0 and all_sign != 0 else True

        # Tier
        tier = "not_recommended"
        reasons = []
        if oc == "diversity": tier = "exploratory"; reasons.append("D&I concentration")
        elif frameworks_agree and cur_all_agree and pos_share > 0.5 and median_est != 0:
            n_supporting = sum(1 for s in [rv_sign, fy_sign, fq_sign] if s == np.sign(median_est))
            if pos_share >= 0.7 and sig5 >= 3: tier = "strong"
            elif pos_share >= 0.6: tier = "moderate"
            elif pos_share >= 0.5: tier = "weak"
            if not frameworks_agree: reasons.append("frameworks disagree on sign")
            if not cur_all_agree: reasons.append("current vs all disagree")
        else:
            if not frameworks_agree: reasons.append("frameworks disagree")
            if not cur_all_agree: reasons.append("current vs all disagree")

        summary_rows.append({
            "outcome":oc,"employee_sample":emp,
            "n_specs":len(sub),"pos_share":pos_share,"neg_share":neg_share,
            "sig5_count":sig5,"sig10_count":sig10,"sig1_count":sig1,
            "median_estimate":median_est,"mean_estimate":mean_est,"iqr":iqr,
            "review_sign":rv_sign,"firm_year_sign":fy_sign,"firm_quarter_sign":fq_sign,
            "frameworks_agree":frameworks_agree,"cur_all_agree":cur_all_agree,
            "tier":tier,"direction":"positive" if median_est>0 else "negative",
            "reasons":"; ".join(reasons) if reasons else "ok",
            "n_events_max":int(sub["n_events"].max()),
            "n_gvkeys_max":int(sub["n_gvkeys"].max()),
        })

df_cons = pd.DataFrame(summary_rows).sort_values(["employee_sample","pos_share"], ascending=[True,False])
df_cons.to_csv(IN_DIR / "focused_rdd_consistency_summary.csv", index=False)

# Best directions
best = df_cons[(df_cons["tier"].isin(["strong","moderate"])) & (df_cons["employee_sample"]=="current")]
df_best = best.sort_values("pos_share", ascending=False)
df_best.to_csv(IN_DIR / "focused_rdd_best_directions.csv", index=False)

print("\n=== Consistency Summary (current employees) ===")
for _, r in df_cons[df_cons["employee_sample"]=="current"].iterrows():
    tier_mark = {"strong":"★★★","moderate":"★★","weak":"★","exploratory":"?","not_recommended":"✗"}.get(r["tier"],"")
    print(f"  {r['outcome']:20s} | {tier_mark:4s} {r['tier']:15s} | pos={r['pos_share']:.0%} | "
          f"median={r['median_estimate']:+.3f} | sig5={int(r['sig5_count'])}/{int(r['n_specs'])} | "
          f"rv={int(r['review_sign']):+d} fy={int(r['firm_year_sign']):+d} fq={int(r['firm_quarter_sign']):+d} | "
          f"NE={int(r['n_events_max'])} NG={int(r['n_gvkeys_max'])}")

# Build Excel workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
hdr_f, title_f, note_f = Font(bold=True,size=10), Font(bold=True,size=13), Font(italic=True,size=9,color="666666")

def style_hdr(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hdr_f) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def style_row(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def auto_w(ws,mn=10,mx=50):
    for col in ws.columns:
        l = get_column_letter(col[0].column)
        ws.column_dimensions[l].width = min(max(max(len(str(c.value or ""))for c in col)+2,mn),mx)

# README
ws = wb.active; ws.title = "README"
ws.cell(row=1,column=1,value="Focused RDD Search Results").font = title_f; ws.merge_cells("A1:D1")
items = [("Date",datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Frameworks","Review-level DiD-RD, Firm-year aggregate RDD, Firm-quarter aggregate RDD"),
    ("SE","gvkey-clustered in all specifications"),
    ("Main threshold","pre>=1 post>=1 (baseline); pre>=3, pre>=5, total>=10 as robustness"),
    ("Main sample","Current employees; all comments as robustness"),
    ("Outcomes","6 broad outcomes; D&I exploratory"),("Bandwidths","Global, |m|<=0.20, |m|<=0.10"),
    ("Consistency rule","Sign must agree across >=2 frameworks AND across current/all")]
r = 3
for k,v in items: ws.cell(row=r,column=1,value=k).font = Font(bold=True); ws.cell(row=r,column=2,value=v); r+=1
auto_w(ws); ws.column_dimensions["B"].width = 80

# Consistency Summary
ws2 = wb.create_sheet("Consistency Summary")
ws2.cell(row=1,column=1,value="Consistency Summary — Current Employees").font = title_f
r = 3
for c,h in enumerate(["Outcome","Tier","Direction","% Positive","Median Est","Sig5/Total","Review","Firm-Year","Firm-Qtr","Frameworks OK","N Events","N gvkeys"],1):
    ws2.cell(row=r,column=c,value=h)
style_hdr(ws2,r,12); r+=1
for _,row in df_cons[df_cons["employee_sample"]=="current"].iterrows():
    vals = [row["outcome"],row["tier"],row["direction"],f"{row['pos_share']:.0%}",f"{row['median_estimate']:.3f}",
            f"{int(row['sig5_count'])}/{int(row['n_specs'])}",f"{int(row["review_sign"]):+d}",f"{int(row["firm_year_sign"]):+d}",
            f"{int(row["firm_quarter_sign"]):+d}","YES" if row["frameworks_agree"] else "NO",
            int(row["n_events_max"]),int(row["n_gvkeys_max"])]
    for c,v in enumerate(vals,1): ws2.cell(row=r,column=c,value=v)
    if row["tier"]=="strong": ws2.cell(row=r,column=2).font = Font(bold=True,color="006600")
    elif row["tier"]=="exploratory": ws2.cell(row=r,column=2).font = Font(color="CC0000")
    style_row(ws2,r,12); r+=1
auto_w(ws2)

# Best Directions
ws3 = wb.create_sheet("Best Directions")
ws3.cell(row=1,column=1,value="Best Candidate Outcomes").font = title_f
r = 3
for c,h in enumerate(["Rank","Outcome","Tier","Direction","Median Est","SE Range","Frameworks Supporting","Current/All OK","Recommendation"],1):
    ws3.cell(row=r,column=c,value=h)
style_hdr(ws3,r,9); r+=1
for i, (_, row) in enumerate(df_best.iterrows()):
    vals = [i+1,row["outcome"],row["tier"],row["direction"],f"{row['median_estimate']:.3f}",
            f"[{row['median_estimate']-row['iqr']/2:.3f},{row['median_estimate']+row['iqr']/2:.3f}]",
            f"RV={int(row["review_sign"]):+d} FY={int(row["firm_year_sign"]):+d} FQ={int(row["firm_quarter_sign"]):+d}",
            "YES" if row["cur_all_agree"] else "NO",
            "Prioritize for paper" if row["tier"]=="strong" else "Include as secondary" if row["tier"]=="moderate" else "Monitor"]
    for c,v in enumerate(vals,1): ws3.cell(row=r,column=c,value=v)
    style_row(ws3,r,9); r+=1
auto_w(ws3)

wb.save(IN_DIR / "focused_rdd_tables.xlsx")
print(f"\nSaved: focused_rdd_tables.xlsx")

# Coauthor report
top = df_best.iloc[0] if len(df_best)>0 else None
rpt = f"""# Focused RDD Search — Coauthor Report

**Date:** {datetime.now().strftime('%Y-%m-%d')}

## What was run
- **3 frameworks**: Review-level DiD-RD, Firm-year delta RDD, Firm-quarter DiD
- **6-7 outcomes**: Overall Rating, Career Opp, Comp & Benefits, Senior Mgmt, WLB, Culture
- **2 employee samples**: Current, All comments
- **3 bandwidths**: Global, |m|<=0.20, |m|<=0.10
- **Multiple thresholds**: pre>=1 post>=1 (baseline), >=3, >=5, total>=10
- **FE variants**: Election FE + year FE, No FE
- **All SE**: gvkey-clustered

## Top Candidates
"""
for _, row in df_best.iterrows():
    rpt += f"- **{row['outcome']}** ({row['tier']}): median tau={row['median_estimate']:+.3f}, "
    rpt += f"{int(row['pos_share']*100)}% positive, "
    rpt += f"RV={int(row["review_sign"]):+d} FY={int(row["firm_year_sign"]):+d} FQ={int(row["firm_quarter_sign"]):+d}\n"

rpt += """
## Recommendations
"""
if top is not None:
    rpt += f"Priority outcome: **{top['outcome']}** — {top['pos_share']:.0%} of specs positive, "
    rpt += "all three frameworks agree on sign.\n"

rpt += """
## Caveats
- Close-election LATE only
- gvkey-clustered SE may be conservative with few clusters at narrow bandwidths
- Multi-election firms: reviews between close elections assigned to nearest
"""

with open(IN_DIR / "focused_rdd_coauthor_report.md","w") as f: f.write(rpt)

print("Done. All outputs in:", IN_DIR)
