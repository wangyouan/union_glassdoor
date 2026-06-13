#!/usr/bin/env python
"""C+D. Relaxed selection + paper tables for v3."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime

PROJ = Path("/data/disk4/workspace/projects/union_glassdoor")
IN = PROJ / "outputs/rdd_rebuild/focused_rdd_search_v3"
IN_V2 = PROJ / "outputs/rdd_rebuild/focused_rdd_search_v2"

print("Loading...")
df = pd.read_csv(IN / "review_level_spline_variants_results.csv")
df_ll = pd.read_csv(IN / "rdrobust_check_results.csv")
# Load v2 aggregate for supporting evidence
df_fy = pd.read_csv(IN_V2 / "matched_firm_year_results.csv") if (IN_V2/"matched_firm_year_results.csv").exists() else pd.DataFrame()
df_fq = pd.read_csv(IN_V2 / "matched_firm_quarter_results.csv") if (IN_V2/"matched_firm_quarter_results.csv").exists() else pd.DataFrame()

OUTCOMES = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
SCREENS = ["pre>=1_post>=1","pre>=3_post>=3","pre>=5_post>=5","total>=10"]
VARIANTS = ["poly1_non_spline","poly1_spline","poly2_non_spline","poly2_spline"]

# ── Selection ──
grid_rows = []
for oc in OUTCOMES:
    for wd in [365, 180, 90]:
        sub = df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&
                 (df["window_days"]==wd)&(df["bandwidth_label"]=="global")]

        # Collect p-values for all 4 variants × 4 screening rules
        p_grid = {}
        tau_grid = {}
        for v in VARIANTS:
            p_grid[v] = {}
            tau_grid[v] = {}
            for scr in SCREENS:
                s = sub[(sub["variant"]==v)&(sub["screening_rule"]==scr)]
                if len(s)>0:
                    p_grid[v][scr] = s["p_value"].values[0]
                    tau_grid[v][scr] = s["estimate"].values[0]
                else:
                    p_grid[v][scr] = np.nan
                    tau_grid[v][scr] = np.nan

        # Baseline (pre>=1) values
        p1_ns = p_grid["poly1_non_spline"].get("pre>=1_post>=1", np.nan)
        p1_s = p_grid["poly1_spline"].get("pre>=1_post>=1", np.nan)
        p2_ns = p_grid["poly2_non_spline"].get("pre>=1_post>=1", np.nan)
        p2_s = p_grid["poly2_spline"].get("pre>=1_post>=1", np.nan)
        t1_ns = tau_grid["poly1_non_spline"].get("pre>=1_post>=1", np.nan)
        t1_s = tau_grid["poly1_spline"].get("pre>=1_post>=1", np.nan)
        t2_ns = tau_grid["poly2_non_spline"].get("pre>=1_post>=1", np.nan)
        t2_s = tau_grid["poly2_spline"].get("pre>=1_post>=1", np.nan)

        # GATE 1: poly1 OR (either spline or non-spline has same sign and p<0.10)
        gate1_pass = False; gate1_variant = ""
        sign1_ns = np.sign(t1_ns); sign1_s = np.sign(t1_s)
        if not np.isnan(p1_ns) and p1_ns<0.10 and not np.isnan(t1_ns):
            gate1_pass = True; gate1_variant = "poly1_non_spline"
        if not np.isnan(p1_s) and p1_s<0.10 and not np.isnan(t1_s):
            if not gate1_pass or p1_s < p1_ns:
                gate1_pass = True; gate1_variant = "poly1_spline"
        # If both pass, use the one with lower p
        if p1_ns<0.10 and p1_s<0.10:
            gate1_variant = "poly1_non_spline" if p1_ns<=p1_s else "poly1_spline"

        ref_sign = sign1_ns if gate1_variant=="poly1_non_spline" else sign1_s
        if gate1_variant=="poly1_non_spline": ref_sign = sign1_ns
        elif gate1_variant=="poly1_spline": ref_sign = sign1_s
        else: ref_sign = sign1_ns if not np.isnan(sign1_ns) else sign1_s

        # GATE 2: poly2 OR (same sign as ref, p<0.10)
        gate2_pass = False; gate2_variant = ""
        for v in ["poly2_non_spline","poly2_spline"]:
            pv = p_grid[v].get("pre>=1_post>=1", np.nan)
            tv = tau_grid[v].get("pre>=1_post>=1", np.nan)
            if not np.isnan(pv) and pv<0.10 and np.sign(tv)==ref_sign:
                gate2_pass = True; gate2_variant = v; break

        # GATE 3: rdrobust same sign, p<0.10 (default bandwidth)
        ll_default = df_ll[(df_ll["outcome"]==oc)&(df_ll["is_default"]==True)]
        gate3_pass = False; ll_tau = np.nan; ll_p = np.nan
        if len(ll_default)>0:
            ll_tau = ll_default["estimate"].values[0]; ll_p = ll_default["p_value"].values[0]
            if not np.isnan(ll_p) and ll_p<0.10 and np.sign(ll_tau)==ref_sign:
                gate3_pass = True

        # GATE 4: same sign across all four screening rules for poly1 and poly2 variants
        signs1 = set(np.sign([tau_grid[v]["pre>=1_post>=1"] for v in ["poly1_non_spline","poly1_spline"] if not np.isnan(tau_grid[v].get("pre>=1_post>=1",np.nan))]))
        signs_all = set()
        for v in VARIANTS:
            for scr in SCREENS:
                tv = tau_grid[v].get(scr, np.nan)
                if not np.isnan(tv): signs_all.add(np.sign(tv))
        gate4_pass = len(signs_all) <= 1  # All same sign

        is_candidate = gate1_pass and gate2_pass and gate3_pass and gate4_pass

        # Count failures
        failures = []
        if not gate1_pass: failures.append("GATE1")
        if not gate2_pass: failures.append("GATE2")
        if not gate3_pass: failures.append("GATE3")
        if not gate4_pass: failures.append("GATE4")

        # Aggregate support
        agg_support = ""
        if len(df_fy)>0:
            fy = df_fy[(df_fy["outcome"]==oc)&(df_fy["employee_sample"]=="current")]
            if len(fy)>0:
                fy_tau = fy["estimate"].median()
                if np.sign(fy_tau)==ref_sign: agg_support += f"FY(τ={fy_tau:+.3f}) "
        if len(df_fq)>0:
            fq = df_fq[(df_fq["outcome"]==oc)&(df_fq["employee_sample"]=="current")]
            if len(fq)>0:
                fq_tau = fq["estimate"].median()
                if np.sign(fq_tau)==ref_sign: agg_support += f"FQ(τ={fq_tau:+.3f})"

        grid_rows.append({
            "outcome":oc,"window_days":wd,"n_failures":len(failures),
            "is_candidate":is_candidate,
            "gate1":gate1_pass,"gate1_variant":gate1_variant,
            "gate2":gate2_pass,"gate2_variant":gate2_variant,
            "gate3":gate3_pass,"gate4":gate4_pass,
            "ref_sign":ref_sign,
            "p1_ns":p1_ns,"p1_s":p1_s,"p2_ns":p2_ns,"p2_s":p2_s,
            "t1_ns":t1_ns,"t1_s":t1_s,"t2_ns":t2_ns,"t2_s":t2_s,
            "ll_tau":ll_tau,"ll_p":ll_p,
            "failed_gates":",".join(failures),
            "agg_support":agg_support.strip(),
            "n_events":int(sub[sub["variant"]=="poly1_spline"]["n_events"].max()) if len(sub)>0 else 0,
        })

df_grid = pd.DataFrame(grid_rows)
df_grid.to_csv(IN / "relaxed_selection_grid.csv", index=False)

candidates = df_grid[df_grid["is_candidate"]]
near_misses = df_grid[(~df_grid["is_candidate"]) & (df_grid["n_failures"]==1)]
candidates.to_csv(IN / "paper_candidates_relaxed.csv", index=False)
near_misses.to_csv(IN / "near_misses_relaxed.csv", index=False)

print(f"\n=== PAPER CANDIDATES ({len(candidates)}) ===")
for _, r in candidates.iterrows():
    print(f"  {r['outcome']:20s} +/-{int(r['window_days'])}d | G1={r['gate1_variant']} G2={r['gate2_variant']} | "
          f"p1ns={r['p1_ns']:.4f} p1s={r['p1_s']:.4f} p2ns={r['p2_ns']:.4f} p2s={r['p2_s']:.4f} | "
          f"ll_p={r['ll_p']:.4f} | agg={r['agg_support']}")

print(f"\n=== NEAR MISSES ({len(near_misses)}) ===")
for _, r in near_misses.iterrows():
    print(f"  {r['outcome']:20s} +/-{int(r['window_days'])}d | FAIL={r['failed_gates']} | "
          f"p1ns={r['p1_ns']:.4f} p1s={r['p1_s']:.4f} p2ns={r['p2_ns']:.4f} p2s={r['p2_s']:.4f} ll_p={r['ll_p']:.4f}")

# ── Paper tables for top candidate ──
if len(candidates)>0:
    best = candidates.iloc[0]
    oc, wd = best["outcome"], int(best["window_days"])

    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    wb = Workbook()
    thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
    hf, tf, nf = Font(bold=True,size=10), Font(bold=True,size=13), Font(italic=True,size=9,color="666666")
    def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hf) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
    def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]

    ws = wb.active; ws.title = "README"
    ws.cell(row=1,column=1,value=f"Focused RDD v3 — {oc} +/-{wd}d").font = tf; ws.merge_cells("A1:E1")
    r=3
    for k,v in [("Candidate",oc),("Window",f"+/-{wd}d"),("Selection","Gates 1-4 all pass")]:
        ws.cell(row=r,column=1,value=k).font = Font(bold=True); ws.cell(row=r,column=2,value=v); r+=1

    # Table 1: Main — 4 variants + rdrobust
    ws1 = wb.create_sheet("Table1 Main")
    ws1.cell(row=1,column=1,value=f"Table 1: {oc} (Current, +/-{wd}d, global, pre>=1)").font = tf
    ws1.merge_cells("A1:F1")
    r=3
    for c, h in enumerate(["", "Poly1 non-spline", "Poly1 spline", "Poly2 non-spline", "Poly2 spline", "rdrobust LL"], 1):
        ws1.cell(row=r,column=c,value=h)
    sh(ws1,r,6); r+=1
    # Win x Post row
    ws1.cell(row=r,column=1,value="Win x Post").font = Font(bold=True)
    for j, v in enumerate(VARIANTS):
        spec = df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&(df["window_days"]==wd)&
                  (df["bandwidth_label"]=="global")&(df["screening_rule"]=="pre>=1_post>=1")&(df["variant"]==v)]
        if len(spec)>0:
            tau, se, pv = spec["estimate"].values[0], spec["standard_error"].values[0], spec["p_value"].values[0]
            sig = "***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
            ws1.cell(row=r,column=2+j,value=f"{tau:.3f}{sig}")
            ws1.cell(row=r+1,column=2+j,value=f"({se:.3f})").font = nf
    # rdrobust
    lld = df_ll[(df_ll["outcome"]==oc)&(df_ll["is_default"]==True)]
    if len(lld)>0:
        tau_ll, se_ll, pv_ll = lld["estimate"].values[0], lld["standard_error"].values[0], lld["p_value"].values[0]
        sig_ll = "***" if pv_ll<0.01 else "**" if pv_ll<0.05 else "*" if pv_ll<0.10 else ""
        ws1.cell(row=r,column=6,value=f"{tau_ll:.3f}{sig_ll}")
        ws1.cell(row=r+1,column=6,value=f"({se_ll:.3f})").font = nf
    sr(ws1,r,6); sr(ws1,r+1,6); r+=2
    # Bottom
    for label, val in [("N reviews",""),("N elections",""),("N firms",""),("Window",f"+/-{wd}d"),("Employee","Current"),("FE","election FE + year FE"),("SE","gvkey-clustered")]:
        ws1.cell(row=r,column=1,value=label).font = Font(italic=True)
        if label=="N reviews":
            spec = df[(df["outcome"]==oc)&(df["employee_sample"]=="current")&(df["window_days"]==wd)&(df["bandwidth_label"]=="global")&(df["screening_rule"]=="pre>=1_post>=1")&(df["variant"]=="poly1_spline")]
            if len(spec)>0:
                ws1.cell(row=r,column=2,value=f"{int(spec['n_reviews'].values[0]):,}")
                ws1.cell(row=r,column=3,value=f"{int(spec['n_events'].values[0])} events")
                ws1.cell(row=r,column=4,value=f"{int(spec['n_gvkeys'].values[0])} firms")
        elif val: ws1.cell(row=r,column=2,value=val)
        r+=1

    wb.save(IN / "focused_v3_paper_tables.xlsx")
    print(f"\nSaved: focused_v3_paper_tables.xlsx")

# ── Coauthor report ──
rpt = f"""# Focused RDD v3 — Coauthor Report (Relaxed Gates)
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## Gate Definitions
- GATE 1 (poly1 OR): Either non-spline or spline poly1 has p<0.10
- GATE 2 (poly2 OR): Either non-spline or spline poly2 has same sign + p<0.10
- GATE 3 (rdrobust): Local-linear default bandwidth same sign + p<0.10
- GATE 4 (screening): All 4 variants × 4 screening rules have consistent sign

## Paper Candidates: {len(candidates)}
"""
for _, r in candidates.iterrows():
    rpt += f"### {r['outcome']} +/-{int(r['window_days'])}d\n"
    rpt += f"- G1={r['gate1_variant']} (p={r['p1_ns']:.4f} ns, {r['p1_s']:.4f} spl)\n"
    rpt += f"- G2={r['gate2_variant']} (p={r['p2_ns']:.4f} ns, {r['p2_s']:.4f} spl)\n"
    rpt += f"- G3 rdrobust: tau={r['ll_tau']:+.4f} p={r['ll_p']:.4f}\n"
    rpt += f"- G4: signs consistent across all rules\n"
    rpt += f"- Aggregate: {r['agg_support']}\n\n"

if len(near_misses)>0:
    rpt += f"## Near Misses ({len(near_misses)})\n"
    for _, r in near_misses.iterrows():
        rpt += f"- {r['outcome']} +/-{int(r['window_days'])}d: {r['failed_gates']}\n"

if len(candidates)==0:
    rpt += "**NO candidates pass all four gates.** See near-misses above.\n"

with open(IN / "focused_v3_coauthor_report.md","w") as f: f.write(rpt)
print("Saved: focused_v3_coauthor_report.md\nDone.")
