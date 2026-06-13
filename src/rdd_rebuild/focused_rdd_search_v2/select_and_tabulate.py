#!/usr/bin/env python
"""C+D. Select robust results and produce paper tables."""

import pandas as pd, numpy as np
from pathlib import Path
from datetime import datetime

PROJ = Path("/data/disk4/workspace/projects/union_glassdoor")
IN_DIR = PROJ / "outputs/rdd_rebuild/focused_rdd_search_v2"
IN_DIR.mkdir(parents=True, exist_ok=True)

print("Loading...")
df_rv = pd.read_csv(IN_DIR / "review_level_poly12_results.csv")
df_fq = pd.read_csv(IN_DIR / "matched_firm_quarter_results.csv")
df_fy = pd.read_csv(IN_DIR / "matched_firm_year_results.csv")

# Focus on election FE specs (primary)
rv_fe = df_rv[df_rv["fixed_effects"]=="election_FE+year_FE"].copy()

# ── Selection logic ──
OUTCOMES_LIST = ["overall_rating","career_opp","comp_benefit","senior_mgmt","wlb","culture"]
BANDWIDTHS = ["global","|m|<=0.20","|m|<=0.10"]
SCREENS = ["pre>=1_post>=1","pre>=3_post>=3","pre>=5_post>=5","total>=10"]

grid_rows = []
for oc in OUTCOMES_LIST:
    for emp in ["current","all"]:
        for bw in BANDWIDTHS:
            for wd in [365,180,90]:
                for poly in [1,2]:
                    # Get p-values across all four screening rules
                    pvals = {}
                    for scr in SCREENS:
                        s = rv_fe[(rv_fe["outcome"]==oc)&(rv_fe["employee_sample"]==emp)&
                                  (rv_fe["bandwidth_label"]==bw)&(rv_fe["window_days"]==wd)&
                                  (rv_fe["screening_rule"]==scr)&(rv_fe["polynomial_order"]==poly)]
                        pvals[scr] = {"tau": s["estimate"].values[0] if len(s)>0 else np.nan,
                                      "se": s["standard_error"].values[0] if len(s)>0 else np.nan,
                                      "p": s["p_value"].values[0] if len(s)>0 else np.nan,
                                      "n": int(s["n_events"].values[0]) if len(s)>0 else 0}

                    # PASS_POLY check (pre>=1 baseline): sign same and p<0.10 in both poly1 and poly2
                    # Actually PASS_POLY means: for THIS poly, p<0.10 and sign defined. The combined check is across poly.

                    # For the grid, record per-poly stats
                    baseline = pvals.get("pre>=1_post>=1", {})
                    grid_rows.append({
                        "outcome":oc,"employee_sample":emp,"bandwidth":bw,"window_days":wd,
                        "polynomial_order":poly,
                        "p_pre1":baseline.get("p",np.nan),"tau_pre1":baseline.get("tau",np.nan),
                        "p_pre3":pvals.get("pre>=3_post>=3",{}).get("p",np.nan),
                        "p_pre5":pvals.get("pre>=5_post>=5",{}).get("p",np.nan),
                        "p_total10":pvals.get("total>=10",{}).get("p",np.nan),
                        "tau_pre3":pvals.get("pre>=3_post>=3",{}).get("tau",np.nan),
                        "tau_pre5":pvals.get("pre>=5_post>=5",{}).get("tau",np.nan),
                        "tau_total10":pvals.get("total>=10",{}).get("tau",np.nan),
                        "n_events_pre1":baseline.get("n",0),
                    })

df_grid = pd.DataFrame(grid_rows)

# ── Paper candidate check ──
def check_cell(outcome, emp, bw, wd):
    """Check if this outcome/sample/bandwidth/window passes all three gates."""
    r1 = df_grid[(df_grid["outcome"]==outcome)&(df_grid["employee_sample"]==emp)&
                 (df_grid["bandwidth"]==bw)&(df_grid["window_days"]==wd)&(df_grid["polynomial_order"]==1)]
    r2 = df_grid[(df_grid["outcome"]==outcome)&(df_grid["employee_sample"]==emp)&
                 (df_grid["bandwidth"]==bw)&(df_grid["window_days"]==wd)&(df_grid["polynomial_order"]==2)]
    if len(r1)==0 or len(r2)==0: return False, {}

    r1=r1.iloc[0]; r2=r2.iloc[0]

    # PASS_POLY: both poly1 and poly2 have same sign and p<0.10 at pre>=1 baseline
    sign1 = np.sign(r1["tau_pre1"]); sign2 = np.sign(r2["tau_pre1"])
    p1_ok = not np.isnan(r1["p_pre1"]) and r1["p_pre1"]<0.10 and not np.isnan(sign1)
    p2_ok = not np.isnan(r2["p_pre1"]) and r2["p_pre1"]<0.10 and not np.isnan(sign2)
    pass_poly = p1_ok and p2_ok and (sign1==sign2)

    # PASS_SCREENS: for BOTH poly1 and poly2, sign identical and p<0.10 across ALL four rules
    scr_cols_p = ["p_pre1","p_pre3","p_pre5","p_total10"]
    scr_cols_t = ["tau_pre1","tau_pre3","tau_pre5","tau_total10"]
    signs1 = set(np.sign([r1[c] for c in scr_cols_t if not np.isnan(r1[c])]))
    signs2 = set(np.sign([r2[c] for c in scr_cols_t if not np.isnan(r2[c])]))
    p_all1_ok = all(not np.isnan(r1[c]) and r1[c]<0.10 for c in scr_cols_p)
    p_all2_ok = all(not np.isnan(r2[c]) and r2[c]<0.10 for c in scr_cols_p)
    pass_screens = (len(signs1)==1 and len(signs2)==1 and signs1==signs2 and p_all1_ok and p_all2_ok)

    # PASS_SCREENS (strong): p<0.05 in all four
    strong = pass_screens and all(r1[c]<0.05 for c in scr_cols_p) and all(r2[c]<0.05 for c in scr_cols_p)

    # PASS_AGGREGATE: at least one aggregate agrees in sign
    agg_agree = False
    agg_info = []
    for agg_df, agg_name in [(df_fq,"firm_quarter"),(df_fy,"firm_year")]:
        a = agg_df[(agg_df["outcome"]==outcome)&(agg_df["employee_sample"]==emp)&
                   (agg_df["bandwidth_label"]==bw)&(agg_df["window_days"]==wd)&
                   (agg_df["screening_rule"]=="pre>=1_post>=1")]
        for poly in [1,2]:
            ap = a[a["polynomial_order"]==poly]
            if len(ap)>0:
                ap_sign = np.sign(ap["estimate"].values[0])
                if ap_sign==sign1:
                    agg_agree = True
                    agg_info.append(f"{agg_name}_poly{poly}: tau={ap['estimate'].values[0]:.3f} p={ap['p_value'].values[0]:.3f}")
    pass_agg = agg_agree

    info = {"sign":sign1,"pass_poly":pass_poly,"pass_screens":pass_screens,"strong":strong,
            "pass_agg":pass_agg,"agg_info":"; ".join(agg_info),
            "p1_pre1":r1["p_pre1"],"p2_pre1":r2["p_pre1"],
            "p1_pre3":r1["p_pre3"],"p2_pre3":r2["p_pre3"],
            "p1_pre5":r1["p_pre5"],"p2_pre5":r2["p_pre5"],
            "p1_total10":r1["p_total10"],"p2_total10":r2["p_total10"],
            "tau1":r1["tau_pre1"],"tau2":r2["tau_pre1"],"n_events":r1["n_events_pre1"]}
    is_candidate = pass_poly and pass_screens and pass_agg
    return is_candidate, info

candidates, near_misses = [], []
for oc in OUTCOMES_LIST:
    for emp in ["current"]:
        for bw in BANDWIDTHS:
            for wd in [365,180,90]:
                ok, info = check_cell(oc, emp, bw, wd)
                row = {"outcome":oc,"employee_sample":emp,"bandwidth":bw,"window_days":wd,**info}
                if ok:
                    candidates.append(row)
                elif info.get("pass_poly") or info.get("pass_screens"):
                    near_misses.append(row)

df_cand = pd.DataFrame(candidates) if candidates else pd.DataFrame()
df_near = pd.DataFrame(near_misses) if near_misses else pd.DataFrame()

df_grid.to_csv(IN_DIR / "robust_selection_grid.csv", index=False)
if len(df_cand)>0: df_cand.to_csv(IN_DIR / "paper_candidates.csv", index=False)
if len(df_near)>0: df_near.to_csv(IN_DIR / "near_misses.csv", index=False)

print(f"\n=== PAPER CANDIDATES ({len(df_cand)}) ===")
if len(df_cand)>0:
    for _, r in df_cand.iterrows():
        print(f"  {r['outcome']:20s} bw={r['bandwidth']:12s} wd={int(r['window_days'])}d | "
              f"p1={r['p1_pre1']:.4f} p2={r['p2_pre1']:.4f} | strong={r['strong']} | NE={int(r['n_events'])}")
        print(f"    screen_pvals: p1=[{r['p1_pre1']:.3f},{r['p1_pre3']:.3f},{r['p1_pre5']:.3f},{r['p1_total10']:.3f}] "
              f"p2=[{r['p2_pre1']:.3f},{r['p2_pre3']:.3f},{r['p2_pre5']:.3f},{r['p2_total10']:.3f}]")
        print(f"    aggregate: {r['agg_info']}")
else:
    print("  NO candidates pass all three gates!")
    print(f"\n=== Near misses ({len(df_near)}) ===")
    for _, r in df_near.iterrows():
        flags = []
        if not r["pass_poly"]: flags.append("FAIL_POLY")
        if not r["pass_screens"]: flags.append("FAIL_SCREENS")
        if not r["pass_agg"]: flags.append("FAIL_AGG")
        print(f"  {r['outcome']:20s} bw={r['bandwidth']:12s} wd={int(r['window_days'])}d | {','.join(flags)} | "
              f"p1_pre1={r['p1_pre1']:.4f} p2_pre1={r['p2_pre1']:.4f}")

# ── Paper tables ──
print("\nBuilding paper tables...")
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
thin = Border(left=Side("thin"),right=Side("thin"),top=Side("thin"),bottom=Side("thin"))
hdr_f, title_f = Font(bold=True,size=10), Font(bold=True,size=13)
note_f = Font(italic=True,size=9,color="666666")
def sh(ws,r,n): [setattr(ws.cell(row=r,column=c),"font",hdr_f) or setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]
def sr(ws,r,n): [setattr(ws.cell(row=r,column=c),"border",thin) for c in range(1,n+1)]

ws = wb.active; ws.title = "README"
ws.cell(row=1,column=1,value="Focused RDD v2 — Paper Tables").font = title_f
ws.merge_cells("A1:D1")
r=3
for k,v in [("Date",datetime.now().strftime("%Y-%m-%d %H:%M")),
    ("Design","Review-level DiD-RD with election FE, poly1 + poly2, 4 screening rules"),
    ("Candidate rule","PASS_POLY + PASS_SCREENS + PASS_AGGREGATE"),
    ("SE","gvkey-clustered")]:
    ws.cell(row=r,column=1,value=k).font = Font(bold=True); ws.cell(row=r,column=2,value=v); r+=1

# Find top candidate
if len(df_cand)>0:
    best_idx = 0
    if "strong" in df_cand.columns:
        strong_mask = df_cand["strong"]==True
        if strong_mask.any():
            best = df_cand[strong_mask].iloc[0]
        else:
            best = df_cand.iloc[0]
    else:
        best = df_cand.iloc[0]
else:
    best = None

if best is not None:
    oc, bw, wd = best["outcome"], best["bandwidth"], int(best["window_days"])

    # Table 1: Main result — all bandwidths, poly1+poly2
    ws1 = wb.create_sheet("Table1 Main Result")
    ws1.cell(row=1,column=1,value=f"Table 1: Review-Level DiD-RD — {oc} (Current, +/-{wd}d, election FE)").font = title_f
    ws1.merge_cells("A1:H1")
    r = 3
    for c, h in enumerate(["","Poly1 Global","Poly1 |m|<=0.20","Poly1 |m|<=0.10","Poly2 Global","Poly2 |m|<=0.20","Poly2 |m|<=0.10"],1):
        ws1.cell(row=r,column=c,value=h)
    sh(ws1,r,7); r+=1
    for scr in SCREENS:
        ws1.cell(row=r,column=1,value=f"Win x Post ({scr})"); ws1.cell(row=r,column=1).font = Font(bold=True)
        for j, poly in enumerate([1,2]):
            for k, bw_l in enumerate(BANDWIDTHS):
                spec = rv_fe[(rv_fe["outcome"]==oc)&(rv_fe["employee_sample"]=="current")&
                             (rv_fe["window_days"]==wd)&(rv_fe["bandwidth_label"]==bw_l)&
                             (rv_fe["screening_rule"]==scr)&(rv_fe["polynomial_order"]==poly)]
                col = 2 + j*3 + k
                if len(spec)>0:
                    tau, se, pv = spec["estimate"].values[0], spec["standard_error"].values[0], spec["p_value"].values[0]
                    sig = "***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
                    ws1.cell(row=r,column=col,value=f"{tau:.3f}{sig}")
                    ws1.cell(row=r+1,column=col,value=f"({se:.3f})").font = note_f
        sr(ws1,r,7); sr(ws1,r+1,7); r+=2
    # Bottom
    r+=1
    for label in ["Window","Employee","FE","SE","Screening rules"]:
        ws1.cell(row=r,column=1,value=label).font = Font(italic=True)
        val = {"Window":f"+/-{wd}d","Employee":"Current","FE":"election FE + year FE","SE":"gvkey-clustered","Screening rules":"all four rules shown"}
        ws1.cell(row=r,column=2,value=val.get(label,"")); r+=1

    # Table 2: Screening robustness — four rules side by side
    ws2 = wb.create_sheet("Table2 Screening Robustness")
    ws2.cell(row=1,column=1,value=f"Table 2: Screening Robustness — {oc} (Current, +/-{wd}d)").font = title_f
    ws2.merge_cells("A1:I1")
    r=3
    for c, h in enumerate(["","pre>=1","","pre>=3","","pre>=5","","total>=10",""],1):
        ws2.cell(row=r,column=c,value=h)
    ws2.cell(row=r+1,column=1,value="Bandwidth")
    for j in range(4): ws2.cell(row=r+1,column=2+j*2,value="Poly1"); ws2.cell(row=r+1,column=3+j*2,value="Poly2")
    sh(ws2,r,9); sh(ws2,r+1,9); r+=2
    for bw_l in BANDWIDTHS:
        ws2.cell(row=r,column=1,value=bw_l)
        for j, scr in enumerate(SCREENS):
            for k, poly in enumerate([1,2]):
                spec = rv_fe[(rv_fe["outcome"]==oc)&(rv_fe["employee_sample"]=="current")&
                             (rv_fe["bandwidth_label"]==bw_l)&(rv_fe["screening_rule"]==scr)&
                             (rv_fe["polynomial_order"]==poly)&(rv_fe["window_days"]==wd)]
                col = 2 + j*2 + k
                if len(spec)>0:
                    tau, pv = spec["estimate"].values[0], spec["p_value"].values[0]
                    sig = "***" if pv<0.01 else "**" if pv<0.05 else "*" if pv<0.10 else ""
                    ws2.cell(row=r,column=col,value=f"{tau:.3f}{sig}")
        sr(ws2,r,9); r+=1

# Save
wb.save(IN_DIR / "focused_v2_paper_tables.xlsx")
print("Saved: focused_v2_paper_tables.xlsx")

# ── Coauthor report ──
rpt = f"""# Focused RDD v2 — Coauthor Report
**Date:** {datetime.now().strftime('%Y-%m-%d')}

## Selection Rule
- PASS_POLY: same sign + p<0.10 in both poly1 AND poly2 (pre>=1 baseline)
- PASS_SCREENS: identical sign + p<0.10 across all four screening rules for BOTH poly1 and poly2
- PASS_AGGREGATE: at least one matched aggregate (firm-quarter or firm-year) agrees in sign

## Results
"""
if len(df_cand)>0:
    rpt += f"**{len(df_cand)} paper candidates found.**\n\n"
    for _, r in df_cand.iterrows():
        rpt += f"### {r['outcome']} | bw={r['bandwidth']} | +/-{int(r['window_days'])}d\n"
        rpt += f"- Poly1 tau={r['tau1']:.3f} (p={r['p1_pre1']:.4f}), Poly2 tau={r['tau2']:.3f} (p={r['p2_pre1']:.4f})\n"
        rpt += f"- Screens: p1=[{r['p1_pre1']:.3f},{r['p1_pre3']:.3f},{r['p1_pre5']:.3f},{r['p1_total10']:.3f}] "
        rpt += f"p2=[{r['p2_pre1']:.3f},{r['p2_pre3']:.3f},{r['p2_pre5']:.3f},{r['p2_total10']:.3f}]\n"
        rpt += f"- Aggregate: {r['agg_info']}\n"
        rpt += f"- Strong (p<0.05 all): {r['strong']}\n"
        rpt += f"- N events: {int(r['n_events'])}\n\n"
else:
    rpt += "**NO candidates pass all three gates.**\n\n"
    if len(df_near)>0:
        rpt += f"### Near misses ({len(df_near)}):\n"
        for _, r in df_near.iterrows():
            rpt += f"- {r['outcome']} bw={r['bandwidth']} +/-{int(r['window_days'])}d: "
            flags = []
            if not r["pass_poly"]: flags.append(f"FAIL_POLY(p1={r['p1_pre1']:.4f},p2={r['p2_pre1']:.4f})")
            if not r["pass_screens"]: flags.append("FAIL_SCREENS")
            if not r["pass_agg"]: flags.append("FAIL_AGG")
            rpt += ", ".join(flags) + "\n"

with open(IN_DIR / "focused_v2_coauthor_report.md","w") as f: f.write(rpt)
print("Saved: focused_v2_coauthor_report.md")
print("Done.")
